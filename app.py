from flask import (
    Flask,
    render_template,
    request,
    redirect,
    flash,
    session,
    jsonify,
    send_file 
)
from flask_sqlalchemy import SQLAlchemy
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, timedelta
import pandas as pd
import unicodedata
import os
import uuid
import json
import geopandas as gpd
from shapely.geometry import Point  
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import openpyxl
import copy
from werkzeug.security import generate_password_hash, check_password_hash


# =========================================
# HORA CDMX
# =========================================

def hora_cdmx():

    return datetime.utcnow() - timedelta(hours=6)

def periodo_quincena():
    ahora = hora_cdmx()
    dia = ahora.day
    mes = ahora.month
    anio = ahora.year

    meses = {
        1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril',
        5:'Mayo', 6:'Junio', 7:'Julio', 8:'Agosto',
        9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'
    }

    import calendar
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    nombre_mes = meses[mes]

    if 10 <= dia <= 14:
        return f"01–15 de {nombre_mes} {anio}"
    elif 25 <= dia <= 29:
        return f"16–{ultimo_dia} de {nombre_mes} {anio}"
    else:
        return f"{nombre_mes} {anio}"

# =========================================
# APP
# =========================================

app = Flask(__name__)


app.secret_key = os.getenv('SECRET_KEY') 
ADMIN_CORREO = os.getenv('ADMIN_CORREO')
ADMIN_CORREO_2 = os.getenv('ADMIN_CORREO_2')
ADMIN_CORREOS = [c for c in [ADMIN_CORREO, ADMIN_CORREO_2] if c]


app.permanent_session_lifetime = timedelta(days=3)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

database_url = os.getenv('DATABASE_URL')

if database_url and database_url.startswith('postgres://'):

    database_url = database_url.replace(
        'postgres://',
        'postgresql://',
        1
    )

app.config['SQLALCHEMY_DATABASE_URI'] = database_url

db = SQLAlchemy(app)

# =========================================
# GEODATA ESTADOS
# =========================================

estados_gdf = gpd.read_file(
    'geodata/estados/dest23gw.shp',
    encoding='latin1'
)

estados_gdf = estados_gdf.to_crs(
    epsg=4326
)

# =========================================
# GEODATA NUCLEOS AGRARIOS
# =========================================

nucleos_gdf = gpd.read_file(
    'geodata/nucleos_agrarios/perimetrales.gpkg',
    encoding='latin1'
)

nucleos_gdf = nucleos_gdf.to_crs(
    epsg=4326
)

# =========================================
# OBTENER UBICACION
# =========================================

def obtener_ubicacion(latitud, longitud):

    punto = Point(
        longitud,
        latitud
    )

    estado = 'No identificado'

    nucleo = 'No identificado'

    estado_resultado = estados_gdf[

        estados_gdf.contains(punto)

    ]

    if not estado_resultado.empty:
        estado = limpiar_texto(

            estado_resultado.iloc[0][
                'NOMGEO'
            ]
        )        

    nucleo_resultado = nucleos_gdf[

        nucleos_gdf.contains(punto)

    ]

    if not nucleo_resultado.empty:

        nucleo = limpiar_texto(

            nucleo_resultado.iloc[0][
                'Name'
            ]
        )

    return {

        'estado': estado,

        'nucleo': nucleo
    }

# =========================================
# MODELO USUARIOS
# =========================================

class Usuario(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    correo = db.Column(
        db.String(200),
        unique=True,
        nullable=False
    )

    nombre = db.Column(
        db.String(200)
    )

    password_hash = db.Column(
        db.String(255)
    )

# =========================================
# SESIONES ACTIVAS
# =========================================

class SesionActiva(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    correo = db.Column(
        db.String(120),
        nullable=False
    )

    token = db.Column(
        db.String(200),
        nullable=False,
        unique=True
    )

    inicio = db.Column(

        db.DateTime,

        default=hora_cdmx
    )

    ultima_actividad = db.Column(

        db.DateTime,

        default=hora_cdmx
    )

# =========================================
# CONTROL DE SESIONES
# =========================================

@app.before_request

def actualizar_sesion():

    token = session.get(
        'session_token'
    )

    if not token:

        return

    sesion_db = SesionActiva.query.filter_by(
        token=token
    ).first()

    # =============================
    # SESIÓN ELIMINADA POR ADMIN
    # =============================

    if not sesion_db:

        session.clear()

        return redirect('/login')

    # =============================
    # ACTUALIZAR ACTIVIDAD
    # =============================

    sesion_db.ultima_actividad = hora_cdmx()

    db.session.commit()

# =========================================
# CONTROL EXPORTACIONES
# =========================================

class Exportacion(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    mes = db.Column(
        db.String(20),
        unique=True,
        nullable=False
    )

    fecha_exportacion = db.Column(

        db.DateTime,

        default=hora_cdmx
    )


# =========================================
# MODELO REGISTROS
# =========================================

class Registro(db.Model):
    
    id = db.Column(
        db.Integer,
        primary_key=True
    )
    usuario = db.Column(db.String(200))

    direccion = db.Column(db.String(200))

    tramo = db.Column(db.String(100))

    entidad = db.Column(db.String(100))

    municipio = db.Column(db.String(100))

    nucleo = db.Column(db.String(200))

    frente = db.Column(db.Integer)

    actividad = db.Column(db.String(100))

    tipo = db.Column(db.String(100))

    mediciones_agroforestales = db.Column(db.Integer)

    mediciones_bdts = db.Column(db.Integer)

    planos = db.Column(db.Integer)

    planos_generados = db.Column(db.Integer)

    planos_validados = db.Column(db.Integer)

    num_infografias = db.Column(db.Integer)

    infografias_generadas = db.Column(db.Integer)

    infografias_validadas = db.Column(db.Integer)

    estatus_infografias = db.Column(db.String(100))

    tipo_propiedad = db.Column(db.String(100))

    observaciones = db.Column(db.Text)

    trabajo_realizado = db.Column(db.String(100))

    actividades_realizadas = db.Column(db.Text)

    trabajo_programado = db.Column(db.String(100))

    actividades_programadas = db.Column(db.Text)

    estatus_trabajo_realizado = db.Column(db.String(100))

    estatus_trabajo_programado = db.Column(db.String(100))

    latitud = db.Column(db.Float)

    longitud = db.Column(db.Float)

    precision_gps = db.Column(db.Float)

    fecha = db.Column(db.DateTime(timezone=True))

# =========================================
# MODELO SUB-ACTIVIDADES
# =========================================

class SubActividad(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    registro_id = db.Column(
        db.Integer,
        db.ForeignKey('registro.id'),
        nullable=False
    )

    tipo = db.Column(
        db.String(20)  # 'realizada' o 'programada'
    )

    entidad = db.Column(db.String(100))

    municipio = db.Column(db.String(100))

    nucleo = db.Column(db.String(200))

    localidad = db.Column(db.String(200))

    frente = db.Column(db.String(20))

    descripcion = db.Column(db.Text)

    trabajo_campo = db.Column(db.String(300))

    # --- Taxonomía del Excel (columnas K–V del REPORTE) ---
    actividad_canonica = db.Column(db.String(60))    # slug de la actividad K–V
    cantidad = db.Column(db.Integer)                 # conteo -> columna Y/TOTAL
    soporte_documental = db.Column(db.String(200))   # columna W (FICHAS, MINUTAS, PLANOS...)

    # =========================================
# CATÁLOGO CANÓNICO DE ACTIVIDADES (columnas K–V del REPORTE)
# slug -> {col Excel, corto (UI), completo (header oficial)}
# =========================================
ACTIVIDADES_CANONICAS = {
    'ASAMBLEAS_COP':             {'col': 'K', 'corto': 'ASISTENCIA A ASAMBLEAS DE LOS NÚCLEOS AGRARIOS INVOLUCRADOS, DE SUSCRIPCIÓN PARA LA SOLICITUD DE ANUENCIA, PARA LA FIRMA DE CONVENIOS DE OCUPACIÓN PREVIA DE USO COMÚN',
        'completo': 'ASISTENCIA A ASAMBLEAS DE LOS NÚCLEOS AGRARIOS INVOLUCRADOS, DE SUSCRIPCIÓN PARA LA SOLICITUD DE ANUENCIA, PARA LA FIRMA DE CONVENIOS DE OCUPACIÓN PREVIA DE USO COMÚN'},
    'REUNIONES_SENSIBILIZACION': {'col': 'L', 'corto': 'ASISTENCIA A REUNIONES INFORMATIVAS DE SENSIBILIZACIÓN CON LOS NÚCLEOS AGRARIOS INVOLUCRADOS',
        'completo': 'ASISTENCIA A REUNIONES INFORMATIVAS DE SENSIBILIZACIÓN CON LOS NÚCLEOS AGRARIOS INVOLUCRADOS'},
    'MEDICION_TOPOGRAFICA':      {'col': 'M', 'corto': 'INSPECCIÓN OCULAR Y TRABAJOS DE MEDICIÓN (LEVANTAMIENTO TOPOGRÁFICO)',
        'completo': 'INSPECCIÓN OCULAR Y TRABAJOS DE MEDICIÓN (LEVANTAMIENTO TOPOGRÁFICO)'},
    'MEDICION_BDT':              {'col': 'N', 'corto': 'INSPECCIÓN OCULAR Y TRABAJOS DE MEDICIÓN EN PROPIEDAD SOCIAL Y PRIVADA PARA LOS BDT´S AGROFORESTALES- CONSTRUCCIÓN',
        'completo': 'INSPECCIÓN OCULAR Y TRABAJOS DE MEDICIÓN EN PROPIEDAD SOCIAL Y PRIVADA PARA LOS BDT´S AGROFORESTALES — CONSTRUCCIÓN'},
    'REVISION_VALIDACION_CAMPO': {'col': 'O', 'corto': 'REVISIÓN Y VALIDACIÓN DE INFORMACIÓN RECOPILADA EN CAMPO POR LA BRIGADA',
        'completo': 'REVISIÓN Y VALIDACIÓN DE INFORMACIÓN RECOPILADA EN CAMPO POR LA BRIGADA'},
    'FICHAS_BDT_CONSTRUCCION':   {'col': 'P', 'corto': 'ELABORACIÓN DE FICHAS DE BIENES DISTINTOS A LA TIERRA DE CONSTRUCCIÓN Y ENVIADAS',
        'completo': 'ELABORACIÓN DE FICHAS DE BIENES DISTINTOS A LA TIERRA DE CONSTRUCCIÓN Y ENVIADAS'},
    'FICHAS_BDT_AGROFORESTAL':   {'col': 'Q', 'corto': 'ELABORACIÓN DE FICHAS DE BIENES DISTINTOS A LA TIERRA AGROFORESTALES Y ENVIADAS',
        'completo': 'ELABORACIÓN DE FICHAS DE BIENES DISTINTOS A LA TIERRA AGROFORESTALES Y ENVIADAS'},
    'PLANOS_CARTOGRAFICOS':      {'col': 'R', 'corto': 'ELABORACIÓN Y VALIDACIÓN DE PLANOS CARTOGRÁFICOS PARA LIBERACIÓN DE DERECHO DE VÍA',
        'completo': 'ELABORACIÓN Y VALIDACIÓN DE PLANOS CARTOGRÁFICOS PARA LIBERACIÓN DE DERECHO DE VÍA'},
    'INFOGRAFIAS_GEOPORTAL':     {'col': 'S', 'corto': 'GENERACIÓN Y/O VALIDACIÓN DE INFOGRAFÍAS, ACTUALIZACIÓN DEL GEOPORTAL',
        'completo': 'GENERACIÓN Y/O VALIDACIÓN DE INFOGRAFÍAS, ACTUALIZACIÓN DEL GEOPORTAL'},
    'CARPETA_BASICA_ASISTENCIA': {'col': 'T', 'corto': 'GESTIÓN DOCUMENTAL DE CARPETA BÁSICA PARA BRINDAR ASISTENCIA TÉCNICA (ANÁLISIS DOCUMENTAL — MEDICIÓN EN CAMPO)',
        'completo': 'GESTIÓN DOCUMENTAL DE CARPETA BÁSICA PARA BRINDAR ASISTENCIA TÉCNICA (ANÁLISIS DOCUMENTAL — MEDICIÓN EN CAMPO)'},
    'TRABAJOS_EXPROPIACION':     {'col': 'U', 'corto': 'ELABORACIÓN DE TRABAJOS TÉCNICOS E INFORMATIVOS DE EXPROPIACIÓN',
        'completo': 'ELABORACIÓN DE TRABAJOS TÉCNICOS E INFORMATIVOS DE EXPROPIACIÓN'},
    'CARPETA_BASICA_AGA':        {'col': 'V', 'corto': 'GESTIÓN DOCUMENTAL DE CARPETA BÁSICA ANTE EL AGA, PARA LA INTEGRACIÓN DE TRABAJOS TÉCNICOS E INFORMATIVOS DE EXPROPIACIÓN',
        'completo': 'GESTIÓN DOCUMENTAL DE CARPETA BÁSICA ANTE EL AGA, PARA LA INTEGRACIÓN DE TRABAJOS TÉCNICOS E INFORMATIVOS DE EXPROPIACIÓN'},
}

# Homologación DIRECCIÓN: valor Xenda -> valor oficial del Excel (para el export)
DIRECCIONES_HOMOLOGACION = {
    'PRODUCTOS Y SISTEMAS GEOESPACIALES':            'PRODUCCIÓN Y SISTEMAS GEOESPACIALES',
    'TOPOGRAFÍA CENTRO':                             'TOPOGRAFIA CENTRO',
    'TOPOGRAFÍA OCCIDENTE':                          'TOPOGRAFIA OCCIDENTE',
    'TOPOGRAFÍA NORTE':                              'TOPOGRAFIA NORTE',
    'TOPOGRAFÍA OCCIDENTE PACÍFICO':                 'TOPOGRAFIA OCCIDENTE PACÍFICO',
    'TOPOGRAFÍA NORTE ORIENTE':                      'TOPOGRAFIA NORTE ORIENTE',
    'TRABAJOS TÉCNICOS INFORMATIVOS DE EXPROPIACIÓN':'TRABAJOS TÉCNICOS INFORMATIVOS DE EXPROPIACIÓN',
}

# =========================================
# REGISTROS ELIMINADOS
# =========================================

class RegistroEliminado(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    id_original = db.Column(db.Integer)

    usuario_original = db.Column(db.String(200))

    eliminado_por = db.Column(db.String(200))

    fecha_eliminacion = db.Column(db.DateTime(timezone=True))

    tramo = db.Column(db.String(100))

    entidad = db.Column(db.String(100))

    municipio = db.Column(db.String(100))

    nucleo = db.Column(db.String(200))

    frente = db.Column(db.Integer)

    actividad = db.Column(db.String(100))

    tipo = db.Column(db.String(100))

    tipo_propiedad = db.Column(db.String(100))

    observaciones = db.Column(db.Text)

    fecha_original = db.Column(db.DateTime(timezone=True))

# =========================================
# NOMBRES DE TRAMOS
# =========================================

TRAMOS_NOMBRES = {
    'TAP':   'AIFA - PACHUCA',
    'TIGDL': 'IRAPUATO - GUADALAJARA',
    'TMLM':  'MAZATLÁN - LOS MOCHIS',
    'TMQ':   'MÉXICO - QUERÉTARO',
    'TQI':   'QUERÉTARO - IRAPUATO',
    'TQSLP': 'QUERÉTARO - SAN LUIS POTOSÍ',
    'TSNL':  'SALTILLO - NUEVO LAREDO',
    'TSLPS': 'SAN LUIS POTOSÍ - SALTILLO',
}

# =========================================
# NORMALIZAR TEXTO
# =========================================

def normalizar(texto):

    texto = str(texto).strip().upper()

    texto = unicodedata.normalize(
        'NFKD',
        texto
    )

    texto = texto.encode(
        'ASCII',
        'ignore'
    ).decode('utf-8')

    return texto

def limpiar_texto(texto):
    
    if not texto:

        return texto

    try:

        return str(texto).encode(
            'latin1'
        ).decode(
            'utf-8'
        )

    except:

        return str(texto)

# =========================================
# CONTROL DE PERIODOS
# =========================================

def registro_habilitado():

    dia = hora_cdmx().day

    return (
        (10 <= dia <= 14)
        or
        (25 <= dia <= 29)
    )


# =========================================
# EXPORTACION AUTOMATICA
# =========================================

def exportar_excel_mensual():

    ahora = hora_cdmx()

    dia = ahora.day

    if dia != 30:

        return

    meses = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
        'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
        'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
        'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    mes_en = ahora.strftime('%B')
    mes_label = f"{meses.get(mes_en, mes_en)} {ahora.year}"
    mes_actual = ahora.strftime('%Y_%m')

    exportado = Exportacion.query.filter_by(
        mes=mes_actual
    ).first()

    if exportado:

        return

    registros = Registro.query.filter(

        db.extract('year', Registro.fecha) == ahora.year,

        db.extract('month', Registro.fecha) == ahora.month

    ).all()

    if not registros:

        return

    datos = []

    for r in registros:

        datos.append({

            'ID': r.id,

            'DIRECCIÓN': r.direccion,

            'FECHA': r.fecha.strftime(
                '%d/%m/%Y %H:%M:%S'
            ) if r.fecha else '',

            'TRAMO': r.tramo,

            'ENTIDAD': r.entidad,

            'MUNICIPIO': r.municipio,

            'NUCLEO': r.nucleo,

            'FRENTE': r.frente,

            'ACTIVIDAD': r.actividad,

            'MODALIDAD': r.tipo,

            'TIPO_PROPIEDAD': r.tipo_propiedad,

            'MEDICIONES_AGROFORESTALES': r.mediciones_agroforestales,

            'MEDICIONES_BDTS': r.mediciones_bdts,

            'PLANOS': r.planos,

            'PLANOS_GENERADOS': r.planos_generados,

            'PLANOS_VALIDADOS': r.planos_validados,

            'NUM_INFOGRAFIAS': r.num_infografias,

            'INFOGRAFIAS_GENERADAS': r.infografias_generadas,

            'INFOGRAFIAS_VALIDADAS': r.infografias_validadas,

            'ESTATUS_INFOGRAFIAS': r.estatus_infografias,

            'TRABAJO_REALIZADO': r.trabajo_realizado,

            'ESTATUS_TRABAJO_REALIZADO': r.estatus_trabajo_realizado,

            'ACTIVIDADES_REALIZADAS': r.actividades_realizadas,

            'TRABAJO_PROGRAMADO': r.trabajo_programado,

            'ESTATUS_TRABAJO_PROGRAMADO': r.estatus_trabajo_programado,

            'ACTIVIDADES_PROGRAMADAS': r.actividades_programadas,

            'OBSERVACIONES': r.observaciones,

            'USUARIO': r.usuario
        })

    df = pd.DataFrame(datos)

    nombre_excel = f'XENDA_{mes_actual}.xlsx'

    ruta_excel = os.path.join(
        os.getcwd(),
        nombre_excel
    )

    df.to_excel(
        ruta_excel,
        index=False
    )

    SCOPES = [
        'https://www.googleapis.com/auth/drive'
    ]

    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if creds_json:
        creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file('credentials.json', scopes=SCOPES)

    service = build(
        'drive',
        'v3',
        credentials=creds
    )

    folders = service.files().list(

        q="mimeType='application/vnd.google-apps.folder' and name='XENDA_REPORTES'",

        spaces='drive',

        fields='files(id, name)'

    ).execute()

    if not folders.get('files'):
        app.logger.warning("Carpeta 'XENDA_REPORTES' no encontrada en Drive; se omite exportación.")
        return

    folder_id = folders['files'][0]['id']

    file_metadata = {

        'name': nombre_excel,

        'parents': [folder_id]

    }

    media = MediaFileUpload(

        ruta_excel,

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    service.files().create(

        body=file_metadata,

        media_body=media,

        fields='id'

    ).execute()

    # =========================================
    # SUBIR REPORTE HTML A DRIVE
    # =========================================

    html_content = generar_reporte_quincenal_html(registros, mes_label)

    nombre_html = f'REPORTE_XENDA_{mes_actual}.html'
    ruta_html = os.path.join(os.getcwd(), nombre_html)

    with open(ruta_html, 'w', encoding='utf-8') as f:
        f.write(html_content)

    file_metadata_html = {
        'name': nombre_html,
        'parents': [folder_id]
    }

    media_html = MediaFileUpload(
        ruta_html,
        mimetype='text/html'
    )

    service.files().create(
        body=file_metadata_html,
        media_body=media_html,
        fields='id'
    ).execute()

    nueva_exportacion = Exportacion(

        mes=mes_actual

    )

    db.session.add(nueva_exportacion)

    db.session.commit()

# =========================================
# GENERAR PRE-REPORTE QUINCENAL HTML
# =========================================

def generar_reporte_quincenal_html(registros, periodo_label):
    quincena = periodo_quincena()

    tramos_nombres = {
        'TAP':   'AIFA - PACHUCA',
        'TIGDL': 'IRAPUATO - GUADALAJARA',
        'TMLM':  'MAZATLÁN - LOS MOCHIS',
        'TMQ':   'MÉXICO - QUERÉTARO',
        'TQI':   'QUERÉTARO - IRAPUATO',
        'TQSLP': 'QUERÉTARO - SAN LUIS POTOSÍ',
        'TSNL':  'SALTILLO - NUEVO LAREDO',
        'TSLPS': 'SAN LUIS POTOSÍ - SALTILLO',
    }

    # Agrupar por dirección, tramo y proceso (Liberación DDV vs Capacitación)
    grupos = {}
    for r in registros:
        proceso = 'CAPACITACIÓN' if (r.actividad or '').strip().upper() == 'CAPACITACIÓN' else 'LIBERACIÓN'
        key = (r.direccion or 'SIN DIRECCIÓN', r.tramo or '', proceso)
        if key not in grupos:
            grupos[key] = []
        grupos[key].append(r)

    secciones_html = ''

    for (direccion, tramo, proceso), regs in sorted(grupos.items()):

        tramo_nombre = tramos_nombres.get(tramo, '') if tramo else ''
        proceso_label = 'Capacitación' if proceso == 'CAPACITACIÓN' else 'Liberación del derecho de vía'

        # Separar por tipo de propiedad
        social = [r for r in regs if r.tipo_propiedad and 'SOCIAL' in r.tipo_propiedad.upper()]
        privada = [r for r in regs if r.tipo_propiedad and 'PRIVADA' in r.tipo_propiedad.upper()]

        # ---- PORTADA DE SECCIÓN ----
        secciones_html += f'''
        <div class="pagina portada-seccion">
            <div class="portada-bandera"></div>
            <div class="portada-contenido">
                <p class="portada-subtitulo">Reporte de actividades</p>
                <div class="portada-divider"></div>
                <p class="portada-periodo">Proyectos Ferroviarios &middot; {quincena}</p>
                <p class="portada-tramo">{'TRAMO ' + tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                <p class="portada-dir">{direccion}</p>
            </div>
        </div>
        '''

        # ---- PROPIEDAD SOCIAL ----
        filas_tabla_social = ''
        if social:

            # SOCIAL — bloques realizados
            bloques_r = ''
            for r in social:
                trabajos_r = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_realizado').all()
                if trabajos_r:
                    for tr in trabajos_r:
                        tipo_tr = tr.frente or ''
                        desc_tr = tr.descripcion or ''
                        estatus_tr = ''
                        if desc_tr.startswith('['):
                            end = desc_tr.find(']')
                            if end > 0:
                                estatus_tr = desc_tr[1:end]
                                desc_tr = desc_tr[end+2:]
                        bloques_r += f'<p><strong>Trabajo de {tipo_tr.lower()}:</strong> <span class="estatus-badge">{estatus_tr}</span></p><p class="acts-texto">{desc_tr.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_realizado:
                    bloques_r += f'<p><strong>Trabajo de {(r.trabajo_realizado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_realizado or ""}</span></p><p class="acts-texto">{(r.actividades_realizadas or "").replace(chr(10), "<br>")}</p>'

            # SOCIAL — bloques programados
            bloques_p = ''
            for r in social:
                trabajos_p = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_programado').all()
                if trabajos_p:
                    for tp in trabajos_p:
                        tipo_tp = tp.frente or ''
                        desc_tp = tp.descripcion or ''
                        estatus_tp = ''
                        if desc_tp.startswith('['):
                            end = desc_tp.find(']')
                            if end > 0:
                                estatus_tp = desc_tp[1:end]
                                desc_tp = desc_tp[end+2:]
                        bloques_p += f'<p><strong>Trabajo de {tipo_tp.lower()}:</strong> <span class="estatus-badge">{estatus_tp}</span></p><p class="acts-texto">{desc_tp.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_programado:
                    bloques_p += f'<p><strong>Trabajo de {(r.trabajo_programado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_programado or ""}</span></p><p class="acts-texto">{(r.actividades_programadas or "").replace(chr(10), "<br>")}</p>'

            secciones_html += f'''
            <div class="pagina">
                <div class="encabezado-pagina">
                    <div class="encabezado-texto">
                        <p class="proyecto">Proyecto ferroviario</p>
                        <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                        <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Social)</span></p>
                    </div>
                    <div class="encabezado-logo"></div>
                </div>
                <div class="seccion-header verde">
                    ACTIVIDADES REALIZADAS EN CAMPO Y/O GABINETE, PROPIEDAD SOCIAL
                </div>
                <div class="seccion-body">{bloques_r}</div>
                <div class="seccion-header guinda">
                    ACTIVIDADES PROGRAMADAS DEL {quincena} EN PROPIEDAD SOCIAL
                </div>
                <div class="seccion-body">{bloques_p or SIN_PROGRAMADAS}</div>
            </div>
            '''
            # ---- TABLA NÚCLEOS SOCIAL ----
            filas_tabla_social = ''
            contador = 1
            for r in social:
                subs = SubActividad.query.filter_by(
                    registro_id=r.id,
                    tipo='realizada'
                ).all()
                if subs:
                    for sub in subs:
                        filas_tabla_social += f'''
                        <tr>
                            <td>{contador}</td>
                            <td>{sub.entidad or ''}</td>
                            <td>{sub.municipio or ''}</td>
                            <td>{sub.nucleo or ''}</td>
                            <td>{('F' + str(sub.frente)) if sub.frente else ''}</td>
                            <td>{'<strong>Trabajo de campo:</strong> ' + sub.trabajo_campo + '<br>' if sub.trabajo_campo else ''}<strong>Actividades:</strong> {(sub.descripcion or '').replace(chr(10), '<br>')}</td>
                        </tr>
                        '''
                        contador += 1
        if filas_tabla_social:
            secciones_html += f'''
            <div class="pagina">
                <div class="encabezado-pagina">
                    <div class="encabezado-texto">
                        <p class="proyecto">Proyecto ferroviario</p>
                        <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                        <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Social)</span></p>
                    </div>
                    <div class="encabezado-logo"></div>
                </div>
                <div class="seccion-header verde">
                    ACTIVIDADES REALIZADAS EN CAMPO (MEDICIÓN) &ndash; PROPIEDAD SOCIAL
                </div>
                <table class="tabla-nucleos">
                    <thead>
                        <tr>
                            <th>No.</th>
                            <th>Entidad Federativa</th>
                            <th>Municipio</th>
                            <th>N&uacute;cleo Agrario</th>
                            <th>Frente</th>
                            <th>Actividades Realizadas</th>
                        </tr>
                    </thead>
                    <tbody>
                        {filas_tabla_social}
                    </tbody>
                </table>
            </div>
            '''
            # TABLA NÚCLEOS PROGRAMADOS SOCIAL
            filas_prog_social = ''
            contador = 1
            for r in social:
                subs = SubActividad.query.filter_by(registro_id=r.id, tipo='programada').all()
                if subs:
                    for sub in subs:
                        filas_prog_social += f'''
                        <tr>
                            <td>{contador}</td>
                            <td>{sub.entidad or ''}</td>
                            <td>{sub.municipio or ''}</td>
                            <td>{sub.nucleo or ''}</td>
                            <td>{('F' + str(sub.frente)) if sub.frente else ''}</td>
                            <td>{'<strong>Trabajo de campo:</strong> ' + sub.trabajo_campo + '<br>' if sub.trabajo_campo else ''}<strong>Actividades:</strong> {(sub.descripcion or '').replace(chr(10), '<br>')}</td>
                        </tr>
                        '''
                        contador += 1

            if filas_prog_social:
                secciones_html += f'''
                <div class="pagina">
                    <div class="encabezado-pagina">
                        <div class="encabezado-texto">
                            <p class="proyecto">Proyecto ferroviario</p>
                            <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                            <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Social)</span></p>
                        </div>
                        <div class="encabezado-logo"></div>
                    </div>
                    <div class="seccion-header guinda">
                        ACTIVIDADES PROGRAMADAS EN CAMPO (MEDICIÓN) &ndash; PROPIEDAD SOCIAL
                    </div>
                    <table class="tabla-nucleos">
                        <thead><tr>
                            <th>No.</th><th>Entidad Federativa</th><th>Municipio</th>
                            <th>N&uacute;cleo Agrario</th><th>Frente</th><th>Actividades Programadas</th>
                        </tr></thead>
                        <tbody>{filas_prog_social}</tbody>
                    </table>
                </div>
                '''

        # ---- PROPIEDAD PRIVADA ----
        filas_tabla_priv = ''
        if privada:

            # Bloques realizados
            bloques_r = ''
            for r in privada:
                trabajos_r = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_realizado').all()
                if trabajos_r:
                    for tr in trabajos_r:
                        tipo_tr = tr.frente or ''
                        desc_tr = tr.descripcion or ''
                        estatus_tr = ''
                        if desc_tr.startswith('['):
                            end = desc_tr.find(']')
                            if end > 0:
                                estatus_tr = desc_tr[1:end]
                                desc_tr = desc_tr[end+2:]
                        bloques_r += f'<p><strong>Trabajo de {tipo_tr.lower()}:</strong> <span class="estatus-badge">{estatus_tr}</span></p><p class="acts-texto">{desc_tr.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_realizado:
                    bloques_r += f'<p><strong>Trabajo de {(r.trabajo_realizado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_realizado or ""}</span></p><p class="acts-texto">{(r.actividades_realizadas or "").replace(chr(10), "<br>")}</p>'

            # Bloques programados
            bloques_p = ''
            for r in privada:
                trabajos_p = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_programado').all()
                if trabajos_p:
                    for tp in trabajos_p:
                        tipo_tp = tp.frente or ''
                        desc_tp = tp.descripcion or ''
                        estatus_tp = ''
                        if desc_tp.startswith('['):
                            end = desc_tp.find(']')
                            if end > 0:
                                estatus_tp = desc_tp[1:end]
                                desc_tp = desc_tp[end+2:]
                        bloques_p += f'<p><strong>Trabajo de {tipo_tp.lower()}:</strong> <span class="estatus-badge">{estatus_tp}</span></p><p class="acts-texto">{desc_tp.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_programado:
                    bloques_p += f'<p><strong>Trabajo de {(r.trabajo_programado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_programado or ""}</span></p><p class="acts-texto">{(r.actividades_programadas or "").replace(chr(10), "<br>")}</p>'

            secciones_html += f'''
            <div class="pagina">
                <div class="encabezado-pagina">
                    <div class="encabezado-texto">
                        <p class="proyecto">Proyecto ferroviario</p>
                        <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                        <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Privada)</span></p>
                    </div>
                    <div class="encabezado-logo"></div>
                </div>
                <div class="seccion-header verde">
                    ACTIVIDADES REALIZADAS EN CAMPO Y/O GABINETE, PROPIEDAD PRIVADA
                </div>
                <div class="seccion-body">{bloques_r}</div>
                <div class="seccion-header guinda">
                    ACTIVIDADES PROGRAMADAS DEL {quincena} EN PROPIEDAD PRIVADA
                </div>
                <div class="seccion-body">{bloques_p or SIN_PROGRAMADAS}</div>
            </div>
            '''
            # ---- TABLA NÚCLEOS PRIVADA ----
            filas_tabla_priv = ''
            contador = 1
            for r in privada:
                subs = SubActividad.query.filter_by(
                    registro_id=r.id,
                    tipo='realizada'
                ).all()
                if subs:
                    for sub in subs:
                        filas_tabla_priv += f'''
                        <tr>
                            <td>{contador}</td>
                            <td>{sub.entidad or ''}</td>
                            <td>{sub.municipio or ''}</td>
                            <td>{sub.localidad or ''}</td>
                            <td>{('F' + str(sub.frente)) if sub.frente else ''}</td>
                            <td>{'<strong>Trabajo de campo:</strong> ' + sub.trabajo_campo + '<br>' if sub.trabajo_campo else ''}<strong>Actividades:</strong> {(sub.descripcion or '').replace(chr(10), '<br>')}</td>
                        </tr>
                        '''
                        contador += 1
        if filas_tabla_priv:
            secciones_html += f'''
            <div class="pagina">
                <div class="encabezado-pagina">
                    <div class="encabezado-texto">
                        <p class="proyecto">Proyecto ferroviario</p>
                        <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                        <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Privada)</span></p>
                    </div>
                    <div class="encabezado-logo"></div>
                </div>
                <div class="seccion-header verde">
                    ACTIVIDADES REALIZADAS EN CAMPO (MEDICIÓN) &ndash; PROPIEDAD PRIVADA
                </div>
                <table class="tabla-nucleos">
                    <thead>
                        <tr>
                            <th>No.</th>
                            <th>Entidad Federativa</th>
                            <th>Municipio</th>
                            <th>Localidad</th>
                            <th>Frente</th>
                            <th>Actividades Realizadas</th>
                        </tr>
                    </thead>
                    <tbody>
                        {filas_tabla_priv}
                    </tbody>
                </table>
            </div>
            '''
             # TABLA NÚCLEOS PROGRAMADOS PRIVADA
            filas_prog_privada = ''
            contador = 1
            for r in privada:
                subs = SubActividad.query.filter_by(registro_id=r.id, tipo='programada').all()
                if subs:
                    for sub in subs:
                        filas_prog_privada += f'''
                        <tr>
                            <td>{contador}</td>
                            <td>{sub.entidad or ''}</td>
                            <td>{sub.municipio or ''}</td>
                            <td>{sub.localidad or ''}</td>
                            <td>{('F' + str(sub.frente)) if sub.frente else ''}</td>
                            <td>{'<strong>Trabajo de campo:</strong> ' + sub.trabajo_campo + '<br>' if sub.trabajo_campo else ''}<strong>Actividades:</strong> {(sub.descripcion or '').replace(chr(10), '<br>')}</td>
                        </tr>
                        '''
                        contador += 1

            if filas_prog_privada:
                secciones_html += f'''
                <div class="pagina">
                    <div class="encabezado-pagina">
                        <div class="encabezado-texto">
                            <p class="proyecto">Proyecto ferroviario</p>
                            <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                            <p class="liberacion">{proceso_label} <span style="color:#6E152E;">(Propiedad Privada)</span></p>
                        </div>
                        <div class="encabezado-logo"></div>
                    </div>
                    <div class="seccion-header guinda">
                        ACTIVIDADES PROGRAMADAS EN CAMPO (MEDICIÓN) &ndash; PROPIEDAD PRIVADA
                    </div>
                    <table class="tabla-nucleos">
                        <thead><tr>
                            <th>No.</th><th>Entidad Federativa</th><th>Municipio</th>
                            <th>Localidad</th><th>Frente</th><th>Actividades Programadas</th>
                        </tr></thead>
                        <tbody>{filas_prog_privada}</tbody>
                    </table>
                </div>
                '''
        # ---- GABINETE / SIN PROPIEDAD ESPECÍFICA ----
        sin_prop = [r for r in regs if not (r.tipo_propiedad and
                    ('SOCIAL' in r.tipo_propiedad.upper() or 'PRIVADA' in r.tipo_propiedad.upper()))]
        if sin_prop:
            bloques_g = ''
            for r in sin_prop:
                trabajos_r = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_realizado').all()
                if trabajos_r:
                    for tr in trabajos_r:
                        tipo_tr = tr.frente or ''
                        desc_tr = tr.descripcion or ''
                        estatus_tr = ''
                        if desc_tr.startswith('['):
                            end = desc_tr.find(']')
                            if end > 0:
                                estatus_tr = desc_tr[1:end]
                                desc_tr = desc_tr[end+2:]
                        bloques_g += f'<p><strong>Trabajo de {tipo_tr.lower()}:</strong> <span class="estatus-badge">{estatus_tr}</span></p><p class="acts-texto">{desc_tr.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_realizado:
                    bloques_g += f'<p><strong>Trabajo de {(r.trabajo_realizado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_realizado or ""}</span></p><p class="acts-texto">{(r.actividades_realizadas or "").replace(chr(10), "<br>")}</p>'

            bloques_gp = ''
            for r in sin_prop:
                trabajos_p = SubActividad.query.filter_by(registro_id=r.id, tipo='trabajo_programado').all()
                if trabajos_p:
                    for tp in trabajos_p:
                        tipo_tp = tp.frente or ''
                        desc_tp = tp.descripcion or ''
                        estatus_tp = ''
                        if desc_tp.startswith('['):
                            end = desc_tp.find(']')
                            if end > 0:
                                estatus_tp = desc_tp[1:end]
                                desc_tp = desc_tp[end+2:]
                        bloques_gp += f'<p><strong>Trabajo de {tipo_tp.lower()}:</strong> <span class="estatus-badge">{estatus_tp}</span></p><p class="acts-texto">{desc_tp.replace(chr(10), "<br>")}</p>'
                elif r.trabajo_programado:
                    bloques_gp += f'<p><strong>Trabajo de {(r.trabajo_programado or "").lower()}:</strong> <span class="estatus-badge">{r.estatus_trabajo_programado or ""}</span></p><p class="acts-texto">{(r.actividades_programadas or "").replace(chr(10), "<br>")}</p>'

            if bloques_g or bloques_gp:
                secciones_html += f'''
                <div class="pagina">
                    <div class="encabezado-pagina">
                        <div class="encabezado-texto">
                            <p class="proyecto">Proyecto ferroviario</p>
                            <p class="tramo-nombre">{tramo_nombre if tramo_nombre else 'DIRECCIÓN DE ' + direccion}</p>
                            <p class="liberacion">{proceso_label}</p>
                        </div>
                        <div class="encabezado-logo"></div>
                    </div>
                    <div class="seccion-header verde">
                        ACTIVIDADES REALIZADAS EN GABINETE
                    </div>
                    <div class="seccion-body">{bloques_g}</div>
                    <div class="seccion-header guinda">
                        ACTIVIDADES PROGRAMADAS DEL {quincena}
                    </div>
                    <div class="seccion-body">{bloques_gp or SIN_PROGRAMADAS}</div>
                </div>
                '''        

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Pre-Reporte Quincenal &middot; {periodo_label}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    background: #f0f0f0;
    color: #1a1a1a;
  }}
  .pagina {{
    width: 960px;
    min-height: 540px;
    background: url('/static/contenido_reporte.png') no-repeat center center;
    background-size: 100% 100%;
    margin: 30px auto;
    padding: 40px 48px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.12);
    position: relative;
  }}
  /* ENCABEZADO */
  .encabezado-pagina {{
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 20px;
  }}
  .encabezado-texto p.proyecto {{
    font-size: 13px;
    color: #BC945A;
    font-style: italic;
    margin-bottom: 2px;
  }}
  .encabezado-texto p.tramo-nombre {{
    font-size: 20px;
    font-weight: bold;
    color: #6E152E;
    text-transform: uppercase;
    margin-bottom: 2px;
  }}
  .encabezado-texto p.liberacion {{
    font-size: 13px;
    color: #245C4F;
    font-weight: 500;
  }}
  .encabezado-logo {{
    width: 120px;
    height: 35px;
    background: url('/static/logo_RAN.png') no-repeat right center;
    background-size: contain;
  }}
  .pagina::after {{
    content: '';
    display: block;
    position: absolute;
    bottom: -15px;
    left: 18px;
    width: 100px;
    height: 90px;
    background: url('/static/gob_mex2-sf.png') no-repeat left center;
    background-size: contain;
  }}
  /* PORTADA */
  .portada-seccion {{
    background: url('/static/portada_reporte.png') no-repeat center center;
    background-size: 100% 100%;
    color: white;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: space-between;
    min-height: 540px;
    padding: 0;
    overflow: hidden;
  }}
  .portada-bandera {{
    position: absolute;
    left: 0;
    bottom: -10px;
    width: 200px;
    height: 260px;
    background: url('/static/bandera.png') no-repeat left center;
    background-size: contain;
    z-index: 10;
  }}
  .portada-seccion::before {{
    content: '';
    position: absolute;
    top: 16px;
    right: 24px;
    width: 380px;
    height: 100px;
    background: url('/static/encabezado_html1.png') no-repeat right center;
    background-size: contain;
    filter: brightness(10);
  }}
  .portada-seccion::after {{
    content: 'Dirección General de Catastro y Asistencia Técnica · Dirección Técnica';
    display: block;
    width: 100%;
    text-align: center;
    font-size: 11px;
    color: #BC945A;
    background: transparent;
    padding: 12px;
    letter-spacing: 0.05em;
  }}
  .portada-contenido {{
    text-align: center;
    padding: 30px 60px;
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
  }}
  .portada-subtitulo {{
    font-size: 15px;
    font-weight: 300;
    margin-bottom: 6px;
    color: #dec9a2;
    letter-spacing: 0.1em;
    text-transform: uppercase;
  }}
  .portada-inst {{
    font-size: 13px;
    color: rgba(222,201,162,0.7);
    margin-bottom: 3px;
  }}
  .portada-divider {{
    width: 60px;
    height: 2px;
    background: #BC945A;
    margin: 20px auto;
  }}
  .portada-periodo {{
    font-size: 15px;
    color: #dec9a2;
    margin-bottom: 16px;
    font-style: italic;
  }}
  .portada-tramo {{
    font-size: 32px;
    font-weight: bold;
    color: white;
    margin-bottom: 10px;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    text-shadow: 0 2px 8px rgba(0,0,0,0.3);
  }}
  .portada-dir {{
    font-size: 15px;
    color: #BC945A;
    text-transform: uppercase;
    font-weight: 600;
    letter-spacing: 0.08em;
    border-top: 1px solid rgba(188,148,90,0.4);
    padding-top: 12px;
    margin-top: 8px;
  }}
  /* SECCIONES */
  .seccion-header {{
    color: #dec9a2;
    font-size: 12px;
    font-weight: bold;
    padding: 10px 16px;
    margin-bottom: 16px;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    text-align: center;
  }}
  .seccion-header.verde {{ background: #245C4F; }}
  .seccion-header.guinda {{ background: #6E152E; margin-top: 24px; }}
  .seccion-body {{
    padding: 0 8px 16px 8px;
    font-size: 13px;
    line-height: 1.7;
  }}
  .seccion-body p {{ margin-bottom: 8px; }}
  .acts-texto {{
    color: #333;
    padding-left: 16px;
    border-left: 3px solid #dec9a2;
    margin-top: 8px;
  }}
  .estatus-badge {{
    display: inline-block;
    background: #dec9a2;
    color: #691B4F;
    font-size: 11px;
    font-weight: bold;
    padding: 2px 10px;
    border-radius: 12px;
    margin-left: 6px;
    text-transform: uppercase;
  }}
  /* TABLA */
  .tabla-nucleos {{
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
    margin-top: 12px;
  }}
  .tabla-nucleos th {{
    background: #245C4F;
    color: #dec9a2;
    padding: 10px 8px;
    text-align: center;
    font-weight: 600;
    text-transform: uppercase;
    font-size: 11px;
    border: 1px solid #1a4438;
  }}
  .tabla-nucleos td {{
    padding: 9px 8px;
    border: 1px solid #d0d0d0;
    vertical-align: middle;
  }}
  .tabla-nucleos tr:nth-child(even) td {{
    background: #f9f6f0;
  }}
  .tabla-nucleos td:nth-child(1) {{ width: 40px; text-align: center; font-weight: bold; }}
  .tabla-nucleos td:nth-child(2) {{ width: 120px; font-weight: bold; text-align: center; }}
  .tabla-nucleos td:nth-child(3) {{ width: 120px; text-align:center; }}
  .tabla-nucleos td:nth-child(4) {{ width: 130px; color: #6E152E; font-weight: bold; text-align: center; }}
  .tabla-nucleos td:nth-child(5) {{ width: 60px; text-align: center; }}
  .tabla-nucleos td:nth-child(6) {{ font-size: 12px; }}
  .tabla-nucleos td {{
    padding: 9px 8px;
    border-bottom: 1px solid #e0e0e0;
    vertical-align: top;
  }}
  .tabla-nucleos tr:nth-child(even) td {{
    background: #f9f6f0;
  }}
  /* PIE */
  @media print {{
    body {{ background: white; }}
    .pagina {{
      box-shadow: none;
      margin: 0;
      page-break-after: always;
    }}
  }}
</style>
</head>
<body>
{secciones_html}
<p style="text-align:center;font-size:11px;color:#aaa;padding:20px;">
  Pre-reporte generado autom&aacute;ticamente por Xenda
</p>
</body>
</html>'''

    return html

# =========================================
# CARGAR CATALOGO
# =========================================

catalogo = pd.read_excel('catalogo_1.xlsx')

catalogo.columns = catalogo.columns.str.strip()

catalogo['TRAMO'] = (
    catalogo['TRAMO']
    .astype(str)
    .str.strip()
)

catalogo['ENTIDAD_FEDERATIVA'] = (
    catalogo['ENTIDAD_FEDERATIVA']
    .astype(str)
    .str.strip()
)

catalogo['MUNICIPIO'] = (
    catalogo['MUNICIPIO']
    .astype(str)
    .str.strip()
)

catalogo['NUCLEO_AGRARIO'] = (
    catalogo['NUCLEO_AGRARIO']
    .astype(str)
    .str.strip()
)

catalogo['ENTIDAD_NORMALIZADA'] = (
    catalogo['ENTIDAD_FEDERATIVA']
    .apply(normalizar)
)

catalogo['MUNICIPIO_NORMALIZADO'] = (
    catalogo['MUNICIPIO']
    .apply(normalizar)
)


# =========================================
# LOGIN
# =========================================

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():

    if request.method == 'POST':

        correo = request.form['correo']

        password = request.form['password']

        usuario = Usuario.query.filter_by(
            correo=correo
        ).first()

        APP_PASSWORD = os.getenv('APP_PASSWORD')

        # Prioridad: hash individual. Fallback: password global (transición, sin lockout)
        if usuario and usuario.password_hash:
            valido = check_password_hash(usuario.password_hash, password)
        elif usuario and APP_PASSWORD:
            valido = (password == APP_PASSWORD)
        else:
            valido = False

        if valido: 

            session.permanent = True

            session['usuario'] = usuario.correo

            # =============================
            # CREAR TOKEN DE SESIÓN
            # =============================

            session_token = str(uuid.uuid4())

            session['session_token'] = session_token

            nueva_sesion = SesionActiva(

                correo=correo,

                token=session_token
            )

            db.session.add(nueva_sesion)

            db.session.commit()

            return redirect('/')

        else:

            return render_template(
                'login.html',
                error='Acceso no autorizado'
            )

    return render_template('login.html')


# =========================================
# LOGOUT
# =========================================

@app.route('/logout')

def logout():

    token = session.get(
        'session_token'
    )

    if token:

        sesion_db = SesionActiva.query.filter_by(
            token=token
        ).first()

        if sesion_db:

            db.session.delete(sesion_db)

            db.session.commit()

    session.clear()

    return redirect('/login')


# =========================================
# ADMIN
# =========================================

@app.route('/admin', methods=['GET', 'POST'])

def admin():

    if session.get('usuario') not in ADMIN_CORREOS:
    
        return redirect('/login')

    if request.method == 'POST':
    
        correo = request.form['correo']

        correo = correo.strip().lower()

        nombre = request.form.get('nombre', '').strip()

        existente = Usuario.query.filter_by(
            correo=correo
        ).first()

        if existente:

            flash('El usuario ya existe')

        else:

            password = request.form.get('password', '').strip()
            nuevo = Usuario(
                correo=correo,
                nombre=nombre or None,
                password_hash=generate_password_hash(password) if password else None
            )

            db.session.add(nuevo)

            db.session.commit()

            flash('Usuario agregado correctamente')

        return redirect('/admin')

    usuarios = Usuario.query.order_by(
        Usuario.correo.asc()
    ).all()

    return render_template(
    'admin.html',
    usuarios=usuarios
)

@app.route('/reset_password/<int:id>', methods=['POST'])
def reset_password(id):
    if session.get('usuario') not in ADMIN_CORREOS:
        return redirect('/login')
    usuario = Usuario.query.get_or_404(id)
    nueva = request.form.get('password', '').strip()
    if not nueva:
        flash('La contraseña no puede estar vacía')
    else:
        usuario.password_hash = generate_password_hash(nueva)
        db.session.commit()
        flash(f'Contraseña actualizada para {usuario.correo}')
    return redirect('/admin')

@app.route('/eliminar_usuario/<int:id>')

def eliminar_usuario(id):

    if 'usuario' not in session:

        return redirect('/login')

    if session['usuario'] not in ADMIN_CORREOS:

        return 'Acceso no autorizado'

    usuario = Usuario.query.get_or_404(id)

    # Protección de correos de administradores para evitar eliminación accidental
    if usuario.correo in ADMIN_CORREOS:

        flash('Este usuario está protegido y no se puede eliminar')

        return redirect('/admin')

    db.session.delete(usuario)

    db.session.commit()

    flash('Usuario eliminado')

    return redirect('/admin')

# =========================================
# REINICIAR REGISTROS
# =========================================

@app.route(
    '/reiniciar_registros',
    methods=['GET', 'POST']
)
def reiniciar_registros():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    def _generar_respaldo():
        regs = Registro.query.all()
        subs = SubActividad.query.all()
        df_reg = pd.DataFrame(
            [{c.name: getattr(r, c.name) for c in Registro.__table__.columns} for r in regs]
        )
        df_sub = pd.DataFrame(
            [{c.name: getattr(s, c.name) for c in SubActividad.__table__.columns} for s in subs]
        )
        nombre = f"XENDA_RESPALDO_{hora_cdmx().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta = os.path.join('/tmp', nombre)
        def _sin_tz(df):
            """Quita zona horaria; Excel no soporta datetimes tz-aware."""
            for col in df.columns:
                s = df[col]
                if isinstance(s.dtype, pd.DatetimeTZDtype):
                    df[col] = s.dt.tz_localize(None)
                else:
                    df[col] = s.map(
                        lambda v: v.replace(tzinfo=None) if hasattr(v, 'tzinfo') and v.tzinfo else v
                    )
            return df

        with pd.ExcelWriter(ruta) as writer:
            _sin_tz(df_reg).to_excel(writer, sheet_name='Registros', index=False)
            _sin_tz(df_sub).to_excel(writer, sheet_name='SubActividades', index=False)
        return ruta, nombre

    def _borrar_todo():
        SubActividad.query.delete()
        Registro.query.delete()
        RegistroEliminado.query.delete()
        Exportacion.query.delete()
        db.session.commit()

    if request.method == 'POST':
        confirmacion = request.form.get('confirmacion', '')
        if confirmacion.upper() != 'CONFIRMAR':
            flash('Escribe CONFIRMAR para continuar')
            return redirect('/reiniciar_registros')

        accion = request.form.get('accion', 'borrar')

        if accion == 'descargar':
            ruta, nombre = _generar_respaldo()   # SIEMPRE antes de borrar
            _borrar_todo()
            return send_file(ruta, as_attachment=True, download_name=nombre)

        _borrar_todo()
        flash('Registros reiniciados correctamente')
        return redirect('/admin')

    return '''
        <h2>¿Seguro que deseas reiniciar TODOS los registros?</h2>
        <p>Esta acción no se puede deshacer.</p>
        <p>Escribe <strong>CONFIRMAR</strong> para continuar:</p>
        <form method="POST">
            <input
                type="text"
                name="confirmacion"
                placeholder="Escribe CONFIRMAR"
                style="padding:8px; font-size:16px; margin:10px 0;"
            >
            <br>
            <button type="submit" name="accion" value="descargar"
                style="padding:10px 16px; margin:6px 4px; background:#2e7d32; color:#fff; border:none; border-radius:6px; cursor:pointer; font-size:15px;">
                Descargar respaldo y borrar
            </button>
            <button type="submit" name="accion" value="borrar"
                style="padding:10px 16px; margin:6px 4px; background:#c62828; color:#fff; border:none; border-radius:6px; cursor:pointer; font-size:15px;">
                Borrar sin descargar
            </button>
            <a href="/admin" style="margin-left:8px;">Cancelar</a>
        </form>
    '''

# =========================================
# SESIONES ACTIVAS
# =========================================

@app.route('/sesiones')

def sesiones():

    if session.get('usuario') not in ADMIN_CORREOS:

        return redirect('/login')

    sesiones_activas = SesionActiva.query.order_by(

        SesionActiva.ultima_actividad.desc()

    ).all()

    return render_template(

        'sesiones.html',

        sesiones=sesiones_activas
    )


# =========================================
# CERRAR SESION REMOTA
# =========================================

@app.route('/cerrar_sesion/<int:id>')

def cerrar_sesion(id):

    if session.get('usuario') not in ADMIN_CORREOS:

        return redirect('/login')

    sesion_obj = SesionActiva.query.get_or_404(id)

    db.session.delete(sesion_obj)

    db.session.commit()

    flash(
        'Sesión cerrada correctamente'
    )

    return redirect('/sesiones')

# =========================================
# CREAR USUARIOS
# =========================================

@app.route('/xenda_admin_bootstrap_users')

def crear_usuarios():

    if session.get('usuario') not in ADMIN_CORREOS:
    
        return redirect('/login')

    correos = [ADMIN_CORREO]

    for correo in correos:

        existente = Usuario.query.filter_by(
            correo=correo
        ).first()

        if not existente:

            nuevo = Usuario(
                correo=correo
            )

            db.session.add(nuevo)

    db.session.commit()

    return 'Usuarios creados correctamente'




# =========================================
# FORMULARIO PRINCIPAL
# =========================================

@app.route('/', methods=['GET', 'POST'])

def index():

#    if (
#        not registro_habilitado()
#        and
#        session.get('usuario') not in ADMIN_CORREOS
#):
#
#        session.clear()
#
#        return render_template(
#            'cerrado.html'
#    )

    if 'usuario' not in session:

        return redirect('/login')

    entidades = sorted(

        catalogo['ENTIDAD_FEDERATIVA']

        .dropna()

        .unique()

    )

    frente = request.form.get('frente')

    frente = int(frente) if frente else None

    if request.method == 'POST':

        nuevo = Registro(
            
            latitud=(
                float(request.form.get('latitud'))
                if request.form.get('latitud')
                else None
            ),

            longitud=(
                float(request.form.get('longitud'))
                if request.form.get('longitud')
                else None
            ),

            precision_gps=(
                float(request.form.get('precision_gps'))
                if request.form.get('precision_gps')
                else None
            ),

            usuario=session['usuario'],

            direccion=request.form.get('direccion'),

            fecha=hora_cdmx(),

            tramo=request.form.get('tramo') or None,

            entidad=request.form.get('entidad') or None,

            municipio=request.form.get('municipio') or None,

            nucleo=request.form.get('nucleo') or None,

            frente=frente,

            actividad=request.form['actividad'],

            tipo=request.form['tipo'],

            mediciones_agroforestales=(request.form.get('mediciones_agroforestales') or 0),

            mediciones_bdts=(request.form.get('mediciones_bdts') or 0),

            planos=(request.form.get('planos') or 0),

            planos_generados=(request.form.get('planos_generados') or 0),

            planos_validados=(request.form.get('planos_validados') or 0),

            num_infografias=(request.form.get('num_infografias') or 0),

            infografias_generadas=(request.form.get('infografias_generadas') or 0),

            infografias_validadas=(request.form.get('infografias_validadas') or 0),

            estatus_infografias=request.form.get('estatus_infografias'),

            tipo_propiedad=request.form.get('tipo_propiedad'),

            observaciones=request.form.get('observaciones', '').upper() or None,

            trabajo_realizado=request.form.get('trabajo_realizado'),

            actividades_realizadas=(
                request.form.get('actividades_realizadas', '').upper()
                or None
            ),

            trabajo_programado=request.form.get('trabajo_programado'),

            actividades_programadas=(
                request.form.get('actividades_programadas', '').upper()
                or None
            ),

            estatus_trabajo_realizado=request.form.get('estatus_trabajo_realizado'),

            estatus_trabajo_programado=request.form.get('estatus_trabajo_programado'),
        )

        db.session.add(nuevo)
        db.session.flush()

        # =====================================
        # GUARDAR SUB-ACTIVIDADES
        # =====================================

        import json

        sub_realizadas = request.form.get('sub_actividades_realizadas', '[]')
        sub_programadas = request.form.get('sub_actividades_programadas', '[]')

        try:
            for item in json.loads(sub_realizadas):
                sub = SubActividad(
                    registro_id=nuevo.id,
                    tipo='realizada',
                    entidad=item.get('entidad', ''),
                    municipio=item.get('municipio', ''),
                    nucleo=item.get('nucleo', ''),
                    localidad=item.get('localidad', ''),
                    frente=item.get('frente', ''),
                    descripcion=item.get('descripcion', ''),
                    trabajo_campo=item.get('trabajo_campo', '')
                )
                db.session.add(sub)
        except:
            pass

        try:
            for item in json.loads(sub_programadas):
                sub = SubActividad(
                    registro_id=nuevo.id,
                    tipo='programada',
                    entidad=item.get('entidad', ''),
                    municipio=item.get('municipio', ''),
                    nucleo=item.get('nucleo', ''),
                    localidad=item.get('localidad', ''),
                    frente=item.get('frente', ''),
                    descripcion=item.get('descripcion', ''),
                    trabajo_campo=item.get('trabajo_campo', '')
                )
                db.session.add(sub)
        except:
            pass

        sub_reporte = request.form.get('sub_actividades_reporte', '[]')
        try:
            for item in json.loads(sub_reporte):
                if not item.get('actividad_canonica'):
                    continue
                sub = SubActividad(
                    registro_id=nuevo.id,
                    tipo='reporte',
                    entidad=item.get('entidad', ''),
                    municipio=item.get('municipio', ''),
                    nucleo=item.get('nucleo', ''),
                    localidad=item.get('localidad', ''),
                    descripcion=item.get('descripcion', ''),
                    actividad_canonica=item.get('actividad_canonica', ''),
                    cantidad=int(item.get('cantidad') or 0) or None,
                    soporte_documental=item.get('soporte_documental', '')
                )
                db.session.add(sub)
        except:
            pass    

        trabajos_realizados_json = request.form.get('trabajos_realizados_json', '[]')
        trabajos_programados_json = request.form.get('trabajos_programados_json', '[]')

        try:
            for item in json.loads(trabajos_realizados_json):
                sub = SubActividad(
                    registro_id=nuevo.id,
                    tipo='trabajo_realizado',
                    entidad='',
                    municipio='',
                    nucleo='',
                    frente=item.get('tipo', ''),
                    descripcion=f"[{item.get('estatus','')}] {item.get('descripcion','')}"
                )
                db.session.add(sub)
        except:
            pass

        try:
            for item in json.loads(trabajos_programados_json):
                sub = SubActividad(
                    registro_id=nuevo.id,
                    tipo='trabajo_programado',
                    entidad='',
                    municipio='',
                    nucleo='',
                    frente=item.get('tipo', ''),
                    descripcion=f"[{item.get('estatus','')}] {item.get('descripcion','')}"
                )
                db.session.add(sub)
        except:
            pass

        db.session.commit()

        flash(
            'Registro guardado exitosamente'
        )

        return redirect('/')

    return render_template(

        'index.html',

        entidades=entidades,

        catalogo_json=catalogo[
            ['TRAMO', 'ENTIDAD_FEDERATIVA', 'MUNICIPIO', 'NUCLEO_AGRARIO']
        ].to_json(orient='records', force_ascii=False)

    )

#=========================================
# ENTIDADES POR TRAMO
#=========================================

@app.route('/entidades/<tramo>')

def entidades_por_tramo(tramo):

    entidades = catalogo[

        catalogo['TRAMO'] == tramo

    ][
        'ENTIDAD_FEDERATIVA'
    ].dropna().unique()

    entidades = sorted(entidades)

    return jsonify(
        list(entidades)
    )


# =========================================
# MUNICIPIOS
# =========================================

@app.route('/municipios/<tramo>/<entidad>')

def municipios(tramo, entidad):

    entidad = normalizar(entidad)

    municipios = catalogo[

        (
            catalogo['TRAMO'] == tramo
        )

        &

        (
            catalogo[
                'ENTIDAD_NORMALIZADA'
            ] == entidad
        )

    ][
        'MUNICIPIO'
    ].dropna().unique()

    municipios = sorted(municipios)

    return jsonify(
        list(municipios)
    )

# =========================================
# NUCLEOS
# =========================================

@app.route('/nucleos/<tramo>/<entidad>/<municipio>')

def nucleos(tramo, entidad, municipio):

    entidad = normalizar(entidad)

    municipio = normalizar(municipio)

    nucleos = catalogo[

        (
            catalogo['TRAMO'] == tramo
        )

        &

        (
            catalogo[
                'ENTIDAD_NORMALIZADA'
            ] == entidad
        )

        &

        (
            catalogo[
                'MUNICIPIO_NORMALIZADO'
            ] == municipio
        )

    ][
        'NUCLEO_AGRARIO'
    ].dropna().unique()

    nucleos = sorted(nucleos)

    return jsonify(
        list(nucleos)
    )

# =========================================
# REGISTROS
# =========================================

@app.route('/registros')

def registros():
    
    if 'usuario' not in session:
    
        return redirect('/login')

    query = Registro.query

    tramo = request.args.get('tramo')

    entidad = request.args.get('entidad')

    municipio = request.args.get('municipio')

    usuario = request.args.get('usuario')

    if tramo:

        query = query.filter(
            Registro.tramo == tramo
        )

    if entidad:

        query = query.filter(
            Registro.entidad == entidad
        )

    if municipio:

        query = query.filter(
            Registro.municipio == municipio
        )

    if usuario:

        query = query.filter(
            Registro.usuario == usuario
        )

    lista = query.order_by(
        Registro.fecha.desc()
    ).all()

    # Desglosar actividades (textarea + bloques + tabla) en filas tipo/estatus/actividad
    def desglosar(reg, subs, kind):
        filas = []
        campo = reg.actividades_realizadas if kind == 'realizado' else reg.actividades_programadas
        tipo_campo = reg.trabajo_realizado if kind == 'realizado' else reg.trabajo_programado
        est_campo = reg.estatus_trabajo_realizado if kind == 'realizado' else reg.estatus_trabajo_programado
        if campo:
            filas.append({'tipo': tipo_campo or '', 'estatus': est_campo or '', 'actividad': campo})
        tb = 'trabajo_' + kind
        for s in subs:
            if s.tipo == tb and s.descripcion:
                desc = s.descripcion
                est = ''
                if desc.startswith('['):
                    fin = desc.find(']')
                    if fin > 0:
                        est = desc[1:fin]
                        desc = desc[fin + 1:].strip()
                filas.append({'tipo': (s.frente or '').strip(), 'estatus': est, 'actividad': desc})
        tt = 'realizada' if kind == 'realizado' else 'programada'
        for s in subs:
            if s.tipo == tt and (s.descripcion or s.trabajo_campo or s.actividad_canonica):
                ubic = ', '.join(x for x in [s.entidad, s.municipio, s.nucleo, s.localidad,
                                             ('F' + str(s.frente)) if s.frente else ''] if x)
                act = s.descripcion or ''
                if s.actividad_canonica:
                    etq = ACTIVIDADES_CANONICAS.get(s.actividad_canonica, {}).get('corto', s.actividad_canonica)
                    qty = f" ×{s.cantidad}" if s.cantidad else ''
                    act = f"[{etq}{qty}] {act}"
                if s.soporte_documental:
                    act = f"{act}  ·  Soporte: {s.soporte_documental}"
                if ubic:
                    act = f"({ubic}) {act}"
                filas.append({'tipo': 'CAMPO', 'estatus': s.trabajo_campo or '', 'actividad': act})
        return filas

    def desglosar_reporte(subs):
        filas = []
        for s in subs:
            if s.tipo == 'reporte' and (s.actividad_canonica or s.descripcion):
                etq = ACTIVIDADES_CANONICAS.get(s.actividad_canonica, {}).get('corto', s.actividad_canonica or '')
                ubic = ', '.join(x for x in [s.entidad, s.municipio, s.nucleo, s.localidad] if x)
                filas.append({
                    'actividad': etq,
                    'cantidad': s.cantidad or '',
                    'soporte': s.soporte_documental or '',
                    'ubicacion': ubic,
                    'descripcion': s.descripcion or ''
                })
        return filas

    ids = [r.id for r in lista]
    subs_all = SubActividad.query.filter(SubActividad.registro_id.in_(ids)).all() if ids else []
    subs_por_reg = {}
    for s in subs_all:
        subs_por_reg.setdefault(s.registro_id, []).append(s)

    for r in lista:
        subs = subs_por_reg.get(r.id, [])
        r.det_realizadas = desglosar(r, subs, 'realizado')
        r.det_programadas = desglosar(r, subs, 'programado')
        r.det_reporte = desglosar_reporte(subs)

    entidades = sorted([
        e[0]
        for e in db.session.query(Registro.entidad).distinct()
        if e[0]
    ])

    municipios = sorted([
        m[0]
        for m in db.session.query(Registro.municipio).distinct()
        if m[0]
    ])

    tramos = sorted([
        t[0]
        for t in db.session.query(Registro.tramo).distinct()
        if t[0]
    ])

    usuarios = sorted([
        u[0]
        for u in db.session.query(Registro.usuario).distinct()
        if u[0]
    ])

    return render_template(

    'registros.html',

    registros=lista,

    admin_correo=ADMIN_CORREO,

    tramos=tramos,

    entidades=entidades,

    municipios=municipios,

    usuarios=usuarios
)

@app.route('/xenda_delete_record/<int:id>')

def eliminar_registro(id):

    if session.get('usuario') not in ADMIN_CORREOS:

        return 'No autorizado', 403

    registro = Registro.query.get_or_404(id)

    SubActividad.query.filter_by(registro_id=id).delete()

    eliminado = RegistroEliminado(

        id_original=registro.id,

        usuario_original=registro.usuario,

        eliminado_por=session['usuario'],

        fecha_eliminacion=hora_cdmx(),
                                       
        tramo=registro.tramo,

        entidad=registro.entidad,

        municipio=registro.municipio,

        nucleo=registro.nucleo,

        frente=registro.frente,

        actividad=registro.actividad,

        tipo=registro.tipo,

        tipo_propiedad=registro.tipo_propiedad,

        observaciones=registro.observaciones,

        fecha_original=registro.fecha
    )

    db.session.add(eliminado)

    db.session.delete(registro)

    db.session.commit()

    return redirect('/registros')

# =========================================
# DESCARGAR USUARIOS
# =========================================

@app.route('/descargar_usuarios')

def descargar_usuarios():

    if session.get('usuario') not in ADMIN_CORREOS:

        return 'No autorizado', 403

    usuarios = Usuario.query.all()

    datos = []

    for u in usuarios:

        datos.append({

            'CORREO': u.correo
        })

    df = pd.DataFrame(datos)

    ruta_archivo = os.path.join(
    '/tmp',
    'usuarios_xenda.xlsx'
    )

    df.to_excel(
        ruta_archivo,
        index=False
    )

    return send_file(
        ruta_archivo,
        as_attachment=True
    )

# =========================================
# DESCARGAR REGISTROS ELIMINADOS
# =========================================

@app.route('/descargar_eliminados')

def descargar_eliminados():

    if session.get('usuario') not in ADMIN_CORREOS:

        return 'No autorizado', 403

    eliminados = RegistroEliminado.query.order_by(
        RegistroEliminado.fecha_eliminacion.desc()
    ).all()

    datos = []

    for r in eliminados:

        datos.append({

            'ID_ORIGINAL': r.id_original,

            'USUARIO_ORIGINAL': r.usuario_original,

            'ELIMINADO_POR': r.eliminado_por,

            'FECHA_ELIMINACION': r.fecha_eliminacion.strftime(
                '%d/%m/%Y %H:%M:%S'
            ) if r.fecha_eliminacion else '',

            'TRAMO': r.tramo,

            'ENTIDAD': r.entidad,

            'MUNICIPIO': r.municipio,

            'NUCLEO': r.nucleo,

            'FRENTE': r.frente,

            'ACTIVIDAD': r.actividad,

            'TIPO': r.tipo,

            'TIPO_PROPIEDAD': r.tipo_propiedad,

            'OBSERVACIONES': r.observaciones,

            'FECHA_ORIGINAL': r.fecha_original.strftime(
                '%d/%m/%Y %H:%M:%S'
            ) if r.fecha_original else '',
        })

    df = pd.DataFrame(datos)

    ruta_archivo = os.path.join(
        '/tmp',
        'registros_eliminados.xlsx'
    )

    df.to_excel(
        ruta_archivo,
        index=False
    )

    return send_file(
        ruta_archivo,
        as_attachment=True
    )

# =========================================
# DESCARGAR REGISTROS 
# =========================================

@app.route('/descargar_registros')
def descargar_registros():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def desglosar_reporte(subs):
        out = []
        for s in subs:
            if s.tipo == 'reporte' and (s.actividad_canonica or s.descripcion):
                etq = ACTIVIDADES_CANONICAS.get(s.actividad_canonica, {}).get('corto', s.actividad_canonica or '')
                out.append({'actividad': etq, 'cantidad': s.cantidad if s.cantidad is not None else '',
                            'soporte': s.soporte_documental or '', 'entidad': s.entidad or '',
                            'municipio': s.municipio or '', 'nucleo': s.nucleo or '',
                            'localidad': s.localidad or '', 'descripcion': s.descripcion or ''})
        return out

    def _trabajos(r, subs, kind):
        out = []
        campo = r.actividades_realizadas if kind == 'realizado' else r.actividades_programadas
        tc = r.trabajo_realizado if kind == 'realizado' else r.trabajo_programado
        ec = r.estatus_trabajo_realizado if kind == 'realizado' else r.estatus_trabajo_programado
        if campo:
            out.append({'actividad': tc or '', 'estatus': ec or '', 'descripcion': campo})
        tb = 'trabajo_' + kind
        for s in subs:
            if s.tipo == tb and s.descripcion:
                desc = s.descripcion; est = ''
                if desc.startswith('['):
                    fin = desc.find(']')
                    if fin > 0:
                        est = desc[1:fin]; desc = desc[fin + 1:].strip()
                out.append({'actividad': (s.frente or '').strip(), 'estatus': est, 'descripcion': desc})
        return out

    def _actividades(subs, kind):
        tt = 'realizada' if kind == 'realizado' else 'programada'
        out = []
        for s in subs:
            if s.tipo == tt and (s.descripcion or s.trabajo_campo):
                out.append({'descripcion': s.descripcion or '', 'entidad': s.entidad or '',
                            'municipio': s.municipio or '', 'nucleo': s.nucleo or '',
                            'localidad': s.localidad or '', 'frente': str(s.frente) if s.frente else '',
                            'trabajo_campo': s.trabajo_campo or ''})
        return out

    SECCIONES = [
        ('', None, [('ID', 'ID')]),
        ('INFORMACIÓN TERRITORIAL', 'DDEBF7', [('Dirección', 'direccion'), ('Tramo', 'tramo'), ('Tipo de Propiedad', 'tipo_propiedad')]),
        ('ACTIVIDAD OPERATIVA', 'E2EFDA', [('Tipo de Actividad', 'actividad'), ('Modalidad', 'tipo'), ('No. de Infografías', 'num_infografias'), ('Infografías Generadas', 'infografias_generadas'), ('Infografías Validadas', 'infografias_validadas'), ('Estatus Infografías', 'estatus_infografias')]),
        ('PRODUCCIÓN TÉCNICA', 'FCE4D6', [('No. de Mediciones', 'mediciones_agroforestales'), ('No. de Fichas', 'mediciones_bdts'), ('Planos', 'planos'), ('Planos Generados', 'planos_generados'), ('Planos Validados', 'planos_validados')]),
        ('REPORTE MATRIZ', 'FFF2CC', [('Actividad', 'rep_actividad'), ('Cantidad', 'rep_cantidad'), ('Soporte', 'rep_soporte'), ('Entidad', 'rep_entidad'), ('Municipio', 'rep_municipio'), ('Núcleo Agrario', 'rep_nucleo'), ('Localidad', 'rep_localidad'), ('Descripción', 'rep_descripcion')]),
        ('TRABAJO REALIZADO', 'D9E1F2', [('Actividad Realizada', 'tr_actividad'), ('Estatus', 'tr_estatus'), ('Descripción', 'tr_descripcion')]),
        ('ACTIVIDADES REALIZADAS (TABLA INTEGRADA)', 'E2EFDA', [('Descripción', 'ar_descripcion'), ('Entidad', 'ar_entidad'), ('Municipio', 'ar_municipio'), ('Núcleo Agrario', 'ar_nucleo'), ('Localidad', 'ar_localidad'), ('Frente', 'ar_frente'), ('Trabajo de Campo', 'ar_trabajo_campo')]),
        ('TRABAJO PROGRAMADO', 'D9E1F2', [('Actividad Programada', 'tp_actividad'), ('Estatus', 'tp_estatus'), ('Descripción', 'tp_descripcion')]),
        ('ACTIVIDADES PROGRAMADAS', 'E2EFDA', [('Descripción', 'ap_descripcion'), ('Entidad', 'ap_entidad'), ('Municipio', 'ap_municipio'), ('Núcleo Agrario', 'ap_nucleo'), ('Localidad', 'ap_localidad'), ('Frente', 'ap_frente'), ('Trabajo de Campo', 'ap_trabajo_campo')]),
        ('REGISTRO', 'D6DCE4', [('Usuario', 'usuario'), ('Fecha', 'fecha')]),
    ]

    query = Registro.query
    tramo = request.args.get('tramo'); entidad = request.args.get('entidad')
    municipio = request.args.get('municipio'); usuario = request.args.get('usuario')
    if tramo:     query = query.filter(Registro.tramo == tramo)
    if entidad:   query = query.filter(Registro.entidad == entidad)
    if municipio: query = query.filter(Registro.municipio == municipio)
    if usuario:   query = query.filter(Registro.usuario == usuario)
    registros = query.order_by(Registro.fecha.desc()).all()

    rows = []
    for r in registros:
        subs = SubActividad.query.filter_by(registro_id=r.id).all()
        rep = desglosar_reporte(subs)
        tr = _trabajos(r, subs, 'realizado'); ar = _actividades(subs, 'realizado')
        tp = _trabajos(r, subs, 'programado'); ap = _actividades(subs, 'programado')
        n = max(len(rep), len(tr), len(ar), len(tp), len(ap), 1)
        for i in range(n):
            f = {'ID': r.id, 'direccion': r.direccion, 'tramo': r.tramo, 'tipo_propiedad': r.tipo_propiedad,
                 'actividad': r.actividad, 'tipo': r.tipo, 'num_infografias': r.num_infografias,
                 'infografias_generadas': r.infografias_generadas, 'infografias_validadas': r.infografias_validadas,
                 'estatus_infografias': r.estatus_infografias, 'mediciones_agroforestales': r.mediciones_agroforestales,
                 'mediciones_bdts': r.mediciones_bdts, 'planos': r.planos, 'planos_generados': r.planos_generados,
                 'planos_validados': r.planos_validados, 'usuario': r.usuario,
                 'fecha': r.fecha.strftime('%d/%m/%Y %H:%M:%S') if r.fecha else ''}
            d = rep[i] if i < len(rep) else {}
            for k in ('actividad', 'cantidad', 'soporte', 'entidad', 'municipio', 'nucleo', 'localidad', 'descripcion'):
                f['rep_' + k] = d.get(k, '')
            d = tr[i] if i < len(tr) else {}
            f['tr_actividad'] = d.get('actividad', ''); f['tr_estatus'] = d.get('estatus', ''); f['tr_descripcion'] = d.get('descripcion', '')
            d = ar[i] if i < len(ar) else {}
            for k in ('descripcion', 'entidad', 'municipio', 'nucleo', 'localidad', 'frente', 'trabajo_campo'):
                f['ar_' + k] = d.get(k, '')
            d = tp[i] if i < len(tp) else {}
            f['tp_actividad'] = d.get('actividad', ''); f['tp_estatus'] = d.get('estatus', ''); f['tp_descripcion'] = d.get('descripcion', '')
            d = ap[i] if i < len(ap) else {}
            for k in ('descripcion', 'entidad', 'municipio', 'nucleo', 'localidad', 'frente', 'trabajo_campo'):
                f['ap_' + k] = d.get(k, '')
            rows.append(f)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Registros'
    thin = Side(style='thin', color='BFBFBF'); borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    COLORES_INST = ['245C4F', '6E152E']
    col = 1
    for idx, (nombre, _color, cols) in enumerate(SECCIONES):
        inst = COLORES_INST[idx % 2]
        relleno = PatternFill('solid', fgColor=inst)
        ini = col
        if not nombre:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            c = ws.cell(row=1, column=col, value=cols[0][0])
            c.font = Font(bold=True, color='FFFFFF'); c.alignment = centro; c.border = borde; c.fill = relleno
            col += 1; continue
        for label, key in cols:
            c2 = ws.cell(row=2, column=col, value=label)
            c2.font = Font(bold=True, size=9, color='FFFFFF'); c2.alignment = centro; c2.border = borde; c2.fill = relleno
            col += 1
        fin = col - 1
        ws.merge_cells(start_row=1, start_column=ini, end_row=1, end_column=fin)
        c1 = ws.cell(row=1, column=ini, value=nombre)
        c1.font = Font(bold=True, size=10, color='FFFFFF'); c1.alignment = centro; c1.border = borde; c1.fill = relleno
    orden = [k for _, _, cols in SECCIONES for _, k in cols]
    for ridx, f in enumerate(rows, start=3):
        for cidx, key in enumerate(orden, start=1):
            c = ws.cell(row=ridx, column=cidx, value=f.get(key, '')); c.alignment = Alignment(vertical='top', wrap_text=True); c.border = borde
    for cidx in range(1, len(orden) + 1):
        ws.column_dimensions[get_column_letter(cidx)].width = 18
    ws.freeze_panes = 'B3'

    ruta_archivo = os.path.join('/tmp', 'registros_xenda.xlsx')
    wb.save(ruta_archivo)
    return send_file(ruta_archivo, as_attachment=True, download_name='registros_xenda.xlsx')

# ============================================================
# EXPORT MATRIZ INFORMES_PROPUESTA (columnas K–V del REPORTE)
# ============================================================
_MATRIZ_COL = {s: v['col'] for s, v in ACTIVIDADES_CANONICAS.items()}   # slug -> letra K..V
PLANTILLA_MATRIZ = os.path.join(os.path.dirname(__file__), 'plantillas', 'INFORMES_PROPUESTA.xlsx')
_MESES_MAY = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
              7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}

def _periodo_excel(anio, mes, q):
    import calendar
    d1, d2 = (1, 15) if q == 1 else (16, calendar.monthrange(anio, mes)[1])
    return f"{d1:02d} AL {d2:02d} DE {_MESES_MAY[mes]} DEL {anio}"

def _prop_excel(v):
    u = (v or '').upper()
    if 'SOCIAL' in u:  return 'PROPIEDAD SOCIAL'
    if 'PRIVADA' in u: return 'PROPIEDAD PRIVADA'
    return u or 'N/A'

def _rellenar_matriz(rows, tpl_path, out_path):
    wb = openpyxl.load_workbook(tpl_path)
    if 'TABLAS DINAMICAS' in wb.sheetnames:          # sin pivotes (las haces a mano)
        del wb['TABLAS DINAMICAS']
    ws = wb['REPORTE']; F0 = 7
    est = {c: (copy.copy(ws.cell(F0,c).font), copy.copy(ws.cell(F0,c).border),
               copy.copy(ws.cell(F0,c).fill), copy.copy(ws.cell(F0,c).alignment),
               ws.cell(F0,c).number_format) for c in range(2, 26)}     # estilos plantilla B..Y
    for r in range(F0, ws.max_row + 1):              # limpia datos viejos
        for c in range(2, 26):
            ws.cell(r, c).value = None
    for i, row in enumerate(rows):
        r = F0 + i
        base = {2:i+1, 3:row['periodo'], 4:row['direccion'], 5:row['tramo'], 6:row['estado'],
                7:row['municipio'], 8:row['nucleo_localidad'] or 'N/A', 9:row['frente'] or 'N/A',
                10:row['tipo_propiedad']}
        for c, v in base.items():
            ws.cell(r, c).value = v
        for c in range(11, 23):                      # K..V = 0
            ws.cell(r, c).value = 0
        col = _MATRIZ_COL.get(row['slug'])
        if col:
            ws[f'{col}{r}'].value = int(row['cantidad'] or 0)
        ws.cell(r, 23).value = row['soporte']         # W
        ws.cell(r, 24).value = row['descripcion']     # X
        ws.cell(r, 25).value = f'=SUM(K{r}:V{r})'     # Y (TOTAL)
        for c in range(2, 26):                        # re-aplica estilos
            f, b, fl, al, nf = est[c]
            cel = ws.cell(r, c)
            cel.font, cel.border, cel.fill, cel.alignment = copy.copy(f), copy.copy(b), copy.copy(fl), copy.copy(al)
            cel.number_format = nf
    ult = F0 + len(rows) - 1 if rows else F0
    ws['AA7'].value = f'=SUM(Y{F0}:Y{max(ult, F0)})'  # ACTIVIDADES TOTALES
    wb.save(out_path)
    return len(rows)


@app.route('/export_matriz')
def export_matriz():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    import calendar
    hoy = hora_cdmx()
    anio = int(request.args.get('anio', hoy.year))
    mes  = int(request.args.get('mes',  hoy.month))
    q    = int(request.args.get('q',    1 if hoy.day <= 15 else 2))

    d1, d2 = (1, 15) if q == 1 else (16, calendar.monthrange(anio, mes)[1])
    ini = datetime(anio, mes, d1, 0, 0, 0)
    fin = datetime(anio, mes, d2, 23, 59, 59)
    periodo_txt = _periodo_excel(anio, mes, q)

    registros = Registro.query.filter(
        Registro.fecha >= ini, Registro.fecha <= fin
    ).order_by(Registro.fecha.asc()).all()

    rows = []
    for reg in registros:
        for s in SubActividad.query.filter_by(registro_id=reg.id, tipo='reporte').all():
            nucleo = (s.nucleo or reg.nucleo or '').strip()
            loc = (s.localidad or '').strip()          # localidad solo existe en la sub
            nuc_loc = f"{nucleo}/{loc}" if (nucleo and loc) else (nucleo or loc or 'N/A')
            rows.append({
                'periodo':          periodo_txt,
                'direccion':        DIRECCIONES_HOMOLOGACION.get(reg.direccion, reg.direccion or ''),
                'tramo':            TRAMOS_NOMBRES.get(reg.tramo, reg.tramo or ''),
                'estado':           s.entidad or reg.entidad or '',
                'municipio':        s.municipio or reg.municipio or '',
                'nucleo_localidad': nuc_loc,
                'frente':           str(reg.frente) if reg.frente is not None else 'N/A',
                'tipo_propiedad':   _prop_excel(reg.tipo_propiedad),
                'slug':             s.actividad_canonica,
                'cantidad':         s.cantidad,
                'soporte':          s.soporte_documental or '',
                'descripcion':      s.descripcion or '',
            })

    out = os.path.join('/tmp', f'REPORTE_MATRIZ_{anio}_{mes:02d}_Q{q}.xlsx')
    _rellenar_matriz(rows, PLANTILLA_MATRIZ, out)
    return send_file(out, as_attachment=True, download_name=os.path.basename(out))

# ============================================================
# EXPORT REPORTE TRIMESTRAL DGCAT (agregado por tramo × actividad)
# ============================================================
PLANTILLA_TRIMESTRAL = os.path.join(os.path.dirname(__file__), 'plantillas', 'TRIMESTRAL_DGCAT.xlsx')

_ACT_KW_TRIM = [   # orden = prioridad (fichas antes que medición-BDT; AGA antes que expropiación)
    ('ASAMBLEAS', 'ASAMBLEAS_COP'),
    ('SENSIBILIZACION', 'REUNIONES_SENSIBILIZACION'),
    ('LEVANTAMIENTO TOPO', 'MEDICION_TOPOGRAFICA'),
    ('CONSTRUCCION Y ENVIADAS', 'FICHAS_BDT_CONSTRUCCION'),
    ('AGROFORESTALES Y ENVIADAS', 'FICHAS_BDT_AGROFORESTAL'),
    ('BDT', 'MEDICION_BDT'),
    ('REVISION Y VALIDACION', 'REVISION_VALIDACION_CAMPO'),
    ('PLANOS CARTOGRAFICOS', 'PLANOS_CARTOGRAFICOS'),
    ('INFOGRAFIAS', 'INFOGRAFIAS_GEOPORTAL'),
    ('ASISTENCIA TECNICA', 'CARPETA_BASICA_ASISTENCIA'),
    ('ANTE EL AGA', 'CARPETA_BASICA_AGA'),
    ('EXPROPIACION', 'TRABAJOS_EXPROPIACION'),
]

def _norm_trim(s):
    s = unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode()
    return s.upper().replace('´', ' ')

def _slug_actividad(b):
    t = _norm_trim(b)
    for kw, slug in _ACT_KW_TRIM:
        if kw in t:
            return slug
    return None

def _tramo_de_celda(a):
    na = _norm_trim(a).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    for code, nombre in TRAMOS_NOMBRES.items():
        key = _norm_trim(nombre).replace(' ', '').replace('-', '')
        if key and key in na:
            return code
    return None

@app.route('/export_trimestral')
def export_trimestral():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    from collections import defaultdict
    hoy = hora_cdmx()
    anio = int(request.args.get('anio', hoy.year))
    trimestre = int(request.args.get('trimestre', (hoy.month - 1) // 3 + 1))
    meses_tri = {1: (1, 2, 3), 2: (4, 5, 6), 3: (7, 8, 9), 4: (10, 11, 12)}
    meses = meses_tri.get(trimestre, (1, 2, 3))

    # --- Agregación (tramo, slug, propiedad) -> suma de cantidad ---
    agg = defaultdict(int)
    registros = Registro.query.filter(
        db.extract('year', Registro.fecha) == anio,
        db.extract('month', Registro.fecha).in_(meses),
    ).all()
    for reg in registros:
        u = (reg.tipo_propiedad or '').upper()
        prop = 'PRIVADA' if 'PRIVADA' in u else 'SOCIAL' if 'SOCIAL' in u else None
        if not reg.tramo or prop is None:
            continue
        for s in SubActividad.query.filter_by(registro_id=reg.id, tipo='reporte').all():
            if s.actividad_canonica and s.cantidad:
                agg[(reg.tramo, s.actividad_canonica, prop)] += int(s.cantidad)

    # --- Rellenar copia de la plantilla (solo D y E) ---
    wb = openpyxl.load_workbook(PLANTILLA_TRIMESTRAL)
    _AUX = {'Coordinador en Campo', 'ANEXO 1'}
    ws = next(wb[s] for s in wb.sheetnames if s not in _AUX)   # la hoja del reporte, se llame como se llame

    MESES_MAY = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
                 7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
    ws['A1'] = f"PERÍODO QUE SE INFORMA: {MESES_MAY[meses[0]]}, {MESES_MAY[meses[1]]} Y {MESES_MAY[meses[2]]}"

    # bloques = rangos entre filas de 'agenda'; tramo = primer tramo hallado en col A del bloque
    agenda_rows = [r for r in range(1, ws.max_row + 1)
                   if 'ATENDER LA AGENDA' in _norm_trim(ws.cell(r, 2).value)]
    agenda_rows.append(ws.max_row + 1)
    for i in range(len(agenda_rows) - 1):
        ini, fin = agenda_rows[i], agenda_rows[i + 1] - 1
        code = next((c for r in range(ini, fin + 1)
                     if (c := _tramo_de_celda(ws.cell(r, 1).value))), None)
        if not code:
            continue
        for r in range(ini, fin + 1):
            slug = _slug_actividad(ws.cell(r, 2).value)
            if not slug:
                continue
            ws.cell(r, 4).value = agg.get((code, slug, 'PRIVADA'), 0) or None
            ws.cell(r, 5).value = agg.get((code, slug, 'SOCIAL'), 0) or None

    out = os.path.join('/tmp', f'REPORTE_TRIMESTRAL_{anio}_T{trimestre}.xlsx')
    wb.save(out)
    return send_file(out, as_attachment=True, download_name=os.path.basename(out))

# =========================================
# DASHBOARD
# =========================================

@app.route('/dashboard')
def dashboard():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403
    return render_template('dashboard.html')

@app.route("/api/dashboard")
def api_dashboard():
    if session.get('usuario') not in ADMIN_CORREOS:
        return jsonify({'error': 'No autorizado'}), 403
    def num(v):
        try: return int(v or 0)
        except (TypeError, ValueError): return 0

    registros = []
    for r in Registro.query.order_by(Registro.fecha.asc()).all():
        try:
            lat = float(r.latitud) if r.latitud not in (None, "") else None
            lng = float(r.longitud) if r.longitud not in (None, "") else None
        except (TypeError, ValueError):
            lat = lng = None
        registros.append({
            "fecha": r.fecha.date().isoformat() if r.fecha else None,
            "direccion": r.direccion or "",
            "tramo": r.tramo or "",
            "propiedad": r.tipo_propiedad or "SIN DATO",
            "actividad": r.actividad or "",
            "num_infografias": num(r.num_infografias),
            "infografias_generadas": num(r.infografias_generadas),
            "infografias_validadas": num(r.infografias_validadas),
            "mediciones_agroforestales": num(r.mediciones_agroforestales),
            "mediciones_bdts": num(r.mediciones_bdts),
            "planos": num(r.planos),
            "planos_generados": num(r.planos_generados),
            "planos_validados": num(r.planos_validados),
            "lat": lat, "lng": lng,
        })
    return jsonify({"registros": registros})

# =========================================
# DASHBOARD DE CONTEO
# =========================================

@app.route('/conteo')
def conteo():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403
    return render_template('conteo.html')


@app.route('/api/conteo')
def api_conteo():
    if session.get('usuario') not in ADMIN_CORREOS:
        return jsonify({'error': 'No autorizado'}), 403

    q = db.session.query(
        Registro.tramo,
        Registro.direccion,
        Registro.tipo_propiedad,
        Registro.actividad,
        db.func.count(Registro.id)
    )

    # Filtros opcionales: ?anio=2026&mes=5  |  ?anio=2026&trimestre=2
    anio = request.args.get('anio', type=int)
    mes = request.args.get('mes', type=int)
    trimestre = request.args.get('trimestre', type=int)
    if anio:
        q = q.filter(db.extract('year', Registro.fecha) == anio)
    if mes:
        q = q.filter(db.extract('month', Registro.fecha) == mes)
    elif trimestre:
        meses_tri = {1: (1, 2, 3), 2: (4, 5, 6), 3: (7, 8, 9), 4: (10, 11, 12)}
        q = q.filter(db.extract('month', Registro.fecha).in_(meses_tri.get(trimestre, (1, 2, 3))))

    q = q.group_by(
        Registro.tramo, Registro.direccion,
        Registro.tipo_propiedad, Registro.actividad
    )

    rows = [
        {
            'tramo': TRAMOS_NOMBRES.get(t, t) if t else 'SIN TRAMO',
            'direccion': d or 'SIN DIRECCIÓN',
            'propiedad': p or 'SIN ESPECIFICAR',
            'actividad': a or 'SIN ACTIVIDAD',
            'n': n
        }
        for t, d, p, a, n in q.all()
    ]
    return jsonify({'rows': rows})



# =========================================
# MAPA GENERAL
# =========================================

@app.route('/mapa_registros')

def mapa_registros():

    if session.get('usuario') not in ADMIN_CORREOS:

        return 'No autorizado', 403

    registros = Registro.query.filter(

        Registro.latitud.isnot(None),

        Registro.longitud.isnot(None)

    ).all()

    for r in registros:
    
        if r.latitud and r.longitud:

            ubicacion = obtener_ubicacion(

                r.latitud,

                r.longitud
            )

            r.estado_geo = ubicacion['estado']

            r.nucleo_geo = ubicacion['nucleo']

        else:

            r.estado_geo = 'Sin coordenadas'

            r.nucleo_geo = 'Sin coordenadas'

    return render_template(

        'mapa_registros.html',

        registros=registros
    )

# =========================================
# MANIFEST PWA
# =========================================

@app.route('/manifest.json')
def manifest():
    return send_file(
        os.path.join(os.path.dirname(__file__), 'manifest.json'),
        mimetype='application/manifest+json'
    )

# =========================================
# SERVICE WORKER PWA
# =========================================

@app.route('/service_worker.js')
def service_worker():
    return send_file('static/service_worker.js', mimetype='application/javascript')

# =========================================
# PRE-REPORTE QUINCENAL
# =========================================

SIN_PROGRAMADAS = '<p class="acts-texto" style="font-weight:bold;">SIN ACTIVIDADES PROGRAMADAS.</p>'

@app.route('/pre_reporte')

def pre_reporte():

    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    ahora = hora_cdmx()

    registros = Registro.query.filter(
        db.extract('year', Registro.fecha) == ahora.year,
        db.extract('month', Registro.fecha) == ahora.month
    ).order_by(Registro.direccion, Registro.tramo).all()

    if not registros:
        return '<h2 style="font-family:sans-serif;padding:40px;">No hay registros en el periodo actual.</h2>'

    meses = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
        'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
        'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
        'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    mes_en = ahora.strftime('%B')
    periodo_label = f"{meses.get(mes_en, mes_en)} {ahora.year}"

    html = generar_reporte_quincenal_html(registros, periodo_label)

    return html

# =========================================
# PRE-REPORTE POR TRAMO
# =========================================

@app.route('/pre_reporte_tramo')
def pre_reporte_tramo():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    ahora = hora_cdmx()
    tramos_nombres = {
        'TAP':   'AIFA - PACHUCA',
        'TIGDL': 'IRAPUATO - GUADALAJARA',
        'TMLM':  'MAZATLÁN - LOS MOCHIS',
        'TMQ':   'MÉXICO - QUERÉTARO',
        'TQI':   'QUERÉTARO - IRAPUATO',
        'TQSLP': 'QUERÉTARO - SAN LUIS POTOSÍ',
        'TSNL':  'SALTILLO - NUEVO LAREDO',
        'TSLPS': 'SAN LUIS POTOSÍ - SALTILLO',
    }

    tramos_disponibles = db.session.query(Registro.tramo).filter(
        db.extract('year', Registro.fecha) == ahora.year,
        db.extract('month', Registro.fecha) == ahora.month,
        Registro.tramo.isnot(None)
    ).distinct().all()

    tramos = [(t[0], tramos_nombres.get(t[0], t[0])) for t in tramos_disponibles]

    return f'''<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Pre-Reporte por Tramo</title>
<style>
  body {{ font-family: Segoe UI, sans-serif; background: #f0f0f0; padding: 40px; }}
  h2 {{ color: #6E152E; margin-bottom: 24px; }}
  .btn {{ display: block; width: 300px; margin: 10px auto; padding: 14px;
          background: #6E152E; color: white; text-align: center;
          border-radius: 10px; text-decoration: none; font-weight: bold; }}
  .btn:hover {{ background: #a42145; }}
</style>
</head>
<body>
<h2 style="text-align:center;">Seleccione un tramo</h2>
{''.join(f'<a class="btn" href="/pre_reporte_tramo/{t[0]}" target="_blank">{t[1]}</a>' for t in sorted(tramos, key=lambda x: x[1]))}
</body></html>'''


@app.route('/pre_reporte_tramo/<tramo>')
def pre_reporte_tramo_detalle(tramo):
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    ahora = hora_cdmx()
    registros = Registro.query.filter(
        db.extract('year', Registro.fecha) == ahora.year,
        db.extract('month', Registro.fecha) == ahora.month,
        Registro.tramo == tramo
    ).order_by(Registro.direccion).all()

    if not registros:
        return '<h2 style="font-family:sans-serif;padding:40px;">No hay registros para este tramo.</h2>'

    meses = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
        'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
        'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
        'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    mes_en = ahora.strftime('%B')
    periodo_label = f"{meses.get(mes_en, mes_en)} {ahora.year}"

    return generar_reporte_quincenal_html(registros, periodo_label)


# =========================================
# PRE-REPORTE POR DIRECCIÓN
# =========================================

@app.route('/pre_reporte_direccion')
def pre_reporte_direccion():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    ahora = hora_cdmx()

    direcciones_disponibles = db.session.query(Registro.direccion).filter(
        db.extract('year', Registro.fecha) == ahora.year,
        db.extract('month', Registro.fecha) == ahora.month
    ).distinct().all()

    direcciones = [d[0] for d in direcciones_disponibles if d[0]]

    return f'''<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Pre-Reporte por Dirección</title>
<style>
  body {{ font-family: Segoe UI, sans-serif; background: #f0f0f0; padding: 40px; }}
  h2 {{ color: #245C4F; margin-bottom: 24px; }}
  .btn {{ display: block; width: 340px; margin: 10px auto; padding: 14px;
          background: #245C4F; color: white; text-align: center;
          border-radius: 10px; text-decoration: none; font-weight: bold; }}
  .btn:hover {{ background: #1a3f36; }}
</style>
</head>
<body>
<h2 style="text-align:center;">Seleccione una dirección</h2>
{''.join(f'<a class="btn" href="/pre_reporte_direccion/{d}" target="_blank">{d}</a>' for d in sorted(direcciones))}
</body></html>'''


@app.route('/pre_reporte_direccion/<direccion>')
def pre_reporte_direccion_detalle(direccion):
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    ahora = hora_cdmx()
    registros = Registro.query.filter(
        db.extract('year', Registro.fecha) == ahora.year,
        db.extract('month', Registro.fecha) == ahora.month,
        Registro.direccion == direccion,
    ).order_by(Registro.fecha).all()

    if not registros:
        return '<h2 style="font-family:sans-serif;padding:40px;">No hay registros para esta dirección.</h2>'

    meses = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
        'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
        'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
        'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    mes_en = ahora.strftime('%B')
    periodo_label = f"{meses.get(mes_en, mes_en)} {ahora.year}"

    return generar_reporte_quincenal_html(registros, periodo_label)

# =========================================
# VERSION
# =========================================

@app.route('/version')
def version():
    return 'v2'

@app.route('/test_html')
def test_html():
    with open('templates/index.html', 'r', encoding='utf-8') as f:
        contenido = f.read()
    if 'CONFIG_DIRECCION' in contenido:
        return 'CONFIG_DIRECCION ENCONTRADO'
    else:
        return 'NO ENCONTRADO'

# =========================================
# CONTEO DETALLADO (quincenal / mensual / trimestral)
# =========================================

MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

@app.route('/conteo_reporte')
def conteo_reporte():
    if session.get('usuario') not in ADMIN_CORREOS:
        return 'No autorizado', 403

    import calendar
    ahora = hora_cdmx()

    periodo   = request.args.get('periodo', 'mensual')
    anio      = request.args.get('anio', type=int) or ahora.year
    mes       = request.args.get('mes', type=int) or ahora.month
    quincena  = request.args.get('quincena', type=int) or (1 if ahora.day <= 15 else 2)
    trimestre = request.args.get('trimestre', type=int) or ((ahora.month - 1) // 3 + 1)

    # ---- Filtro por periodo (usa extract, igual que el resto del código) ----
    q = Registro.query.filter(db.extract('year', Registro.fecha) == anio)
    if periodo == 'trimestral':
        meses_tri = {1:(1,2,3), 2:(4,5,6), 3:(7,8,9), 4:(10,11,12)}
        q = q.filter(db.extract('month', Registro.fecha).in_(meses_tri.get(trimestre, (1,2,3))))
        etiqueta = f"{trimestre}º Trimestre {anio}"
    elif periodo == 'quincenal':
        q = q.filter(db.extract('month', Registro.fecha) == mes)
        if quincena == 1:
            q = q.filter(db.extract('day', Registro.fecha) <= 15)
            etiqueta = f"01–15 de {MESES_ES.get(mes, mes)} {anio}"
        else:
            ult = calendar.monthrange(anio, mes)[1]
            q = q.filter(db.extract('day', Registro.fecha) >= 16)
            etiqueta = f"16–{ult} de {MESES_ES.get(mes, mes)} {anio}"
    else:
        periodo = 'mensual'
        q = q.filter(db.extract('month', Registro.fecha) == mes)
        etiqueta = f"{MESES_ES.get(mes, mes)} {anio}"

    registros = q.all()

    # ---- Sub-actividades de esos registros (1 sola query) ----
    ids = [r.id for r in registros]
    subs = SubActividad.query.filter(SubActividad.registro_id.in_(ids)).all() if ids else []
    sub_por_registro = {}
    for s in subs:
        sub_por_registro.setdefault(s.registro_id, []).append(s)

    PLACEHOLDERS = {'', 'VARIOS', 'PROPIEDAD PRIVADA', 'N/A'}

    def metricas(regs):
        m = dict(registros=len(regs), mediciones=0, fichas=0, planos=0, planos_gen=0,
                 planos_val=0, infografias=0, info_gen=0, info_val=0, nucleos=0)
        nucleos = set()
        for r in regs:
            m['mediciones']  += int(r.mediciones_agroforestales or 0)
            m['fichas']      += int(r.mediciones_bdts or 0)
            m['planos']      += int(r.planos or 0)
            m['planos_gen']  += int(r.planos_generados or 0)
            m['planos_val']  += int(r.planos_validados or 0)
            m['infografias'] += int(r.num_infografias or 0)
            m['info_gen']    += int(r.infografias_generadas or 0)
            m['info_val']    += int(r.infografias_validadas or 0)
            for s in sub_por_registro.get(r.id, []):
                if s.tipo == 'realizada' and s.nucleo and s.nucleo.strip().upper() not in PLACEHOLDERS:
                    nucleos.add(s.nucleo.strip().upper())
        m['nucleos'] = len(nucleos)
        return m

    glob = metricas(registros)

    por_dir = {}
    por_tramo = {}
    for r in registros:
        por_dir.setdefault(r.direccion or 'SIN DIRECCIÓN', []).append(r)
        por_tramo.setdefault(r.tramo or 'SIN TRAMO', []).append(r)

    def celda_gv(total, g, v):
        return f"{total} <span style='color:#999;font-size:11px;'>({g}/{v})</span>" if (g or v) else str(total)

    def fila(nombre, m, i):
        return f'''<tr>
            <td style="text-align:center;">{i}</td>
            <td style="font-weight:600;color:#6E152E;">{nombre}</td>
            <td style="text-align:center;">{m['registros']}</td>
            <td style="text-align:center;">{m['nucleos']}</td>
            <td style="text-align:center;">{m['mediciones']}</td>
            <td style="text-align:center;">{m['fichas']}</td>
            <td style="text-align:center;">{celda_gv(m['planos'], m['planos_gen'], m['planos_val'])}</td>
            <td style="text-align:center;">{celda_gv(m['infografias'], m['info_gen'], m['info_val'])}</td>
        </tr>'''

    def fila_total(m):
        return f'''<tr style="background:#dec9a2;font-weight:bold;">
            <td></td><td style="color:#6E152E;">TOTAL</td>
            <td style="text-align:center;">{m['registros']}</td>
            <td style="text-align:center;">{m['nucleos']}</td>
            <td style="text-align:center;">{m['mediciones']}</td>
            <td style="text-align:center;">{m['fichas']}</td>
            <td style="text-align:center;">{celda_gv(m['planos'], m['planos_gen'], m['planos_val'])}</td>
            <td style="text-align:center;">{celda_gv(m['infografias'], m['info_gen'], m['info_val'])}</td>
        </tr>'''

    filas_dir = ''.join(fila(d, metricas(rs), i+1)
                        for i, (d, rs) in enumerate(sorted(por_dir.items()))) + fila_total(glob)
    filas_tramo = ''.join(fila(TRAMOS_NOMBRES.get(t, t), metricas(rs), i+1)
                          for i, (t, rs) in enumerate(sorted(por_tramo.items()))) + fila_total(glob)

    def kpi(valor, label):
        return f'<div class="kpi"><div class="kpi-num">{valor}</div><div class="kpi-lbl">{label}</div></div>'

    kpis = (kpi(glob['registros'], 'Registros') + kpi(glob['nucleos'], 'Núcleos atendidos') +
            kpi(glob['mediciones'], 'Mediciones') + kpi(glob['fichas'], 'Fichas') +
            kpi(glob['planos'], 'Planos') + kpi(glob['infografias'], 'Infografías'))

    def sel(nombre, opciones, actual):
        ops = ''.join(f'<option value="{v}"{" selected" if str(v)==str(actual) else ""}>{txt}</option>'
                      for v, txt in opciones)
        return f'<select name="{nombre}">{ops}</select>'

    sel_periodo = sel('periodo', [('quincenal','Quincenal'),('mensual','Mensual'),('trimestral','Trimestral')], periodo)
    sel_mes = sel('mes', [(k, v) for k, v in MESES_ES.items()], mes)
    sel_quin = sel('quincena', [(1,'1ª quincena'),(2,'2ª quincena')], quincena)
    sel_tri = sel('trimestre', [(1,'1º'),(2,'2º'),(3,'3º'),(4,'4º')], trimestre)

    return f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Conteo detallado · {etiqueta}</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Segoe UI',Arial,sans-serif; background:#f0f0f0; color:#1a1a1a; padding:30px 20px; }}
  .wrap {{ max-width:1080px; margin:auto; }}
  .head {{ background:#6E152E; color:#dec9a2; border-radius:16px 16px 0 0; padding:22px 28px; }}
  .head h1 {{ font-size:22px; color:white; letter-spacing:2px; }}
  .head p {{ font-size:14px; margin-top:4px; }}
  .filtro {{ background:white; padding:16px 28px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; border-bottom:2px solid #f0e6ec; }}
  .filtro select, .filtro input {{ padding:9px 12px; border-radius:10px; border:1.5px solid #d1d5db; font-size:14px; }}
  .filtro button {{ padding:9px 20px; border:none; border-radius:10px; background:#245C4F; color:white; font-weight:bold; cursor:pointer; }}
  .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:12px; padding:24px 28px; background:white; }}
  .kpi {{ background:#f9f6f0; border-left:4px solid #BC945A; border-radius:10px; padding:16px; text-align:center; }}
  .kpi-num {{ font-size:30px; font-weight:bold; color:#6E152E; }}
  .kpi-lbl {{ font-size:12px; color:#555; text-transform:uppercase; letter-spacing:0.4px; margin-top:4px; }}
  .bloque {{ background:white; padding:8px 28px 28px; }}
  .bloque h2 {{ font-size:13px; color:#245C4F; text-transform:uppercase; letter-spacing:1px; margin:20px 0 12px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th {{ background:#245C4F; color:#dec9a2; padding:10px 8px; text-transform:uppercase; font-size:11px; border:1px solid #1a4438; }}
  td {{ padding:9px 8px; border:1px solid #e0e0e0; }}
  tr:nth-child(even) td {{ background:#f9f6f0; }}
  .foot {{ background:white; border-radius:0 0 16px 16px; padding:16px 28px; font-size:11px; color:#aaa; text-align:center; }}
  .foot a {{ color:#6E152E; text-decoration:none; font-weight:bold; }}
</style></head>
<body><div class="wrap">
  <div class="head">
    <h1>CONTEO DETALLADO</h1>
    <p>Proyectos Ferroviarios · {etiqueta}</p>
  </div>
  <form method="get" class="filtro">
    {sel_periodo}
    <input type="number" name="anio" value="{anio}" style="width:90px;">
    {sel_mes} {sel_quin} {sel_tri}
    <button type="submit">Ver</button>
  </form>
  <div class="kpis">{kpis}</div>
  <div class="bloque">
    <h2>Por Dirección</h2>
    <table>
      <thead><tr><th>No.</th><th>Dirección</th><th>Registros</th><th>Núcleos</th>
        <th>Mediciones</th><th>Fichas</th><th>Planos (gen/val)</th><th>Infografías (gen/val)</th></tr></thead>
      <tbody>{filas_dir}</tbody>
    </table>
    <h2>Por Tramo</h2>
    <table>
      <thead><tr><th>No.</th><th>Tramo</th><th>Registros</th><th>Núcleos</th>
        <th>Mediciones</th><th>Fichas</th><th>Planos (gen/val)</th><th>Infografías (gen/val)</th></tr></thead>
      <tbody>{filas_tramo}</tbody>
    </table>
  </div>
  <div class="foot">Generado por Xenda · <a href="/admin">Volver al panel</a></div>
</div></body></html>'''

# =========================================
# CREAR TABLAS
# =========================================

with app.app_context():
    
    db.create_all()
    # --- Auto-migración: columnas nuevas de sub_actividad (PostgreSQL) ---
    try:
        with db.engine.begin() as conn:
            conn.execute(db.text("ALTER TABLE sub_actividad ADD COLUMN IF NOT EXISTS actividad_canonica VARCHAR(60)"))
            conn.execute(db.text("ALTER TABLE sub_actividad ADD COLUMN IF NOT EXISTS cantidad INTEGER"))
            conn.execute(db.text("ALTER TABLE sub_actividad ADD COLUMN IF NOT EXISTS soporte_documental VARCHAR(200)"))
            conn.execute(db.text("ALTER TABLE usuario ADD COLUMN IF NOT EXISTS password_hash VARCHAR(255)"))
    except Exception as _e:
        print('Auto-migración sub_actividad:', _e)  

# =========================================
# INICIO
# =========================================

if __name__ == '__main__':

    app.run(
        host='0.0.0.0',
        port=5000,
        debug=os.getenv('FLASK_DEBUG') == '1'
    )
